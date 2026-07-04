import json
import subprocess
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "tools" / "build_published_content.py"


class BuildPublishedContentTests(unittest.TestCase):
    def test_builds_runtime_artifacts_from_realistic_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            tmp_path = Path(tmp_dir)
            glossary_path = tmp_path / "glossary.xlsx"
            structure_path = tmp_path / "structure.xlsx"
            out_dir = tmp_path / "published"

            write_glossary_workbook(glossary_path)
            write_structure_workbook(structure_path)

            subprocess.run(
                [
                    str(REPO_ROOT / ".venv" / "bin" / "python3"),
                    str(SCRIPT_PATH),
                    "--glossary-workbook",
                    str(glossary_path),
                    "--structure-workbook",
                    str(structure_path),
                    "--out-dir",
                    str(out_dir),
                    # This fixture exercises the workbook + auto-classification
                    # tiers in isolation, so disable the authoritative registry
                    # (which would otherwise match these real-world term titles
                    # and mask the per-source breakdown this test pins).
                    "--taxonomy-registry",
                    "__disabled_registry_for_fixture__",
                ],
                check=True,
            )

            index_payload = json.loads((out_dir / "terms" / "index.json").read_text(encoding="utf-8"))
            manifest_payload = json.loads((out_dir / "terms" / "manifest.json").read_text(encoding="utf-8"))
            paths_payload = json.loads((out_dir / "paths" / "index.json").read_text(encoding="utf-8"))
            report_payload = json.loads((out_dir / "reports" / "import-report.json").read_text(encoding="utf-8"))
            audit_payload = json.loads((out_dir / "reports" / "content-audit.json").read_text(encoding="utf-8"))
            manifest_payload = json.loads((out_dir / "manifest.json").read_text(encoding="utf-8"))
            structure_payload = json.loads(
                (out_dir / "editorial" / "structure-registry.json").read_text(encoding="utf-8")
            )
            launch_contract = json.loads((out_dir / "editorial" / "launch-contract.json").read_text(encoding="utf-8"))
            term_summary = next(item for item in index_payload if item["slug"] == "activation-function")
            shard_payload = json.loads(
                (out_dir / "terms" / "shards" / f"{term_summary['artifact']['shardId']}.json").read_text(encoding="utf-8")
            )
            term_payload = next(item for item in shard_payload["terms"] if item["slug"] == "activation-function")
            path_payload = json.loads(
                (out_dir / "paths" / "by-slug" / "neural-networks-advanced-architectures.json").read_text(encoding="utf-8")
            )

            transformer_summary = next(item for item in index_payload if item["slug"] == "transformer")
            vision_transformer_summary = next(item for item in index_payload if item["slug"] == "vision-transformer")
            rag_summary = next(item for item in index_payload if item["slug"] == "retrieval-augmented-generation-rag")
            near_duplicate_summary = next(item for item in index_payload if item["slug"] == "near-duplicate-detection-minhash")
            act_r_summary = next(item for item in index_payload if item["slug"] == "act-r-architecture")
            act_r_shard = json.loads(
                (out_dir / "terms" / "shards" / f"{act_r_summary['artifact']['shardId']}.json").read_text(
                    encoding="utf-8"
                )
            )
            act_r_term = next(item for item in act_r_shard["terms"] if item["slug"] == "act-r-architecture")

            self.assertEqual(len(index_payload), 9)
            self.assertEqual(report_payload["sourceInventoryTermCount"], 10)
            self.assertEqual(manifest_payload["termCount"], 9)
            self.assertEqual(report_payload["termCount"], 9)
            # Tiered taxonomy model: the workbook contributes 5 explicit
            # classifications (columns N/O/P rows), auto-classification rules
            # add more on top. taxonomyMatches is the union, so it is a floor
            # (>= workbookClassifiedTerms), not an exact total. The per-source
            # breakdown fields pin each tier independently.
            self.assertEqual(report_payload["workbookClassifiedTerms"], 5)
            self.assertGreaterEqual(report_payload["taxonomyMatches"], report_payload["workbookClassifiedTerms"])
            self.assertEqual(
                report_payload["taxonomyMatches"],
                report_payload["workbookClassifiedTerms"]
                + report_payload["autoClassifiedTerms"]
                + report_payload["registryClassifiedTerms"]
                + report_payload["studyFamilyClassifiedTerms"],
            )
            self.assertEqual(report_payload["definitionMatches"], 1)
            self.assertEqual(report_payload["canonicalizationGroups"], 1)
            self.assertEqual(report_payload["canonicalizationRowsMerged"], 1)
            self.assertEqual(report_payload["pathCount"], 1)
            self.assertGreaterEqual(report_payload["shardCount"], 1)
            self.assertGreater(structure_payload["fieldCount"], 2)
            self.assertGreater(structure_payload["layerCounts"]["launch-runtime"], 0)
            self.assertGreater(structure_payload["layerCounts"]["editorial-expansion"], 0)
            self.assertGreater(structure_payload["layerCounts"]["backlog"], 0)
            self.assertIn("Introduction", structure_payload["launchSections"])
            self.assertIn("Case Studies", structure_payload["editorialSections"])
            self.assertEqual(report_payload["structureLayerCounts"], structure_payload["layerCounts"])
            self.assertEqual(report_payload["structureSectionCount"], len(structure_payload["sectionGroups"]))
            self.assertEqual(report_payload["launchStructureSections"], len(structure_payload["launchSections"]))
            self.assertEqual(report_payload["editorialStructureSections"], len(structure_payload["editorialSections"]))
            self.assertEqual(
                structure_payload["editorialSections"][:3],
                ["Historical Context", "Case Studies", "Hands-on Tutorials"],
            )
            self.assertEqual(manifest_payload["termCount"], 9)
            self.assertEqual(manifest_payload["pathCount"], 1)
            self.assertEqual(manifest_payload["launchSectionCount"], 11)
            self.assertEqual(manifest_payload["coverage"]["blockCoverage"]["study-prompts"], 9)
            self.assertEqual(manifest_payload["contentDepth"]["blockCounts"]["structure-expansion"], 9)
            self.assertEqual(manifest_payload["contentDepth"]["blockCounts"]["curriculum-map"], 9)
            self.assertEqual(manifest_payload["structureLayerCounts"], structure_payload["layerCounts"])
            self.assertEqual(manifest_payload["launchSections"], launch_contract["launchSections"])
            self.assertEqual(launch_contract["launchSectionCount"], 11)
            self.assertEqual(launch_contract["sheetName"], "Sheet2")
            self.assertIn("overview", launch_contract["launchBlockIds"])
            self.assertEqual(launch_contract["sourceStructureCounts"], structure_payload["layerCounts"])
            self.assertEqual(
                [item["section"] for item in launch_contract["launchSections"]],
                list(structure_payload["launchSections"]),
            )
            self.assertEqual(audit_payload["status"], "pass")
            self.assertEqual(audit_payload["termCount"], 9)
            self.assertEqual(audit_payload["qualityChecks"]["highSeverityIssueCount"], 0)
            self.assertEqual(audit_payload["coverage"]["blockCoverage"]["overview"], 9)
            self.assertEqual(audit_payload["coverage"]["blockCoverage"]["study-prompts"], 9)
            self.assertEqual(audit_payload["coverage"]["sourceDefinitionBlocks"], 1)
            self.assertEqual(term_payload["taxonomy"]["category"], "Neural Networks")
            self.assertEqual(term_payload["source"]["glossaryWorkbook"]["definitionRow"], 3)
            self.assertEqual(term_payload["aliases"], ["Activation-Function"])
            self.assertIn("B", act_r_term["source"]["glossaryWorkbook"]["inventoryColumns"])
            self.assertIn("Neural Networks / Components", term_payload["summary"])
            self.assertIn("Connection Map", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Why It Matters", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Comparison Notes", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Recall Drill", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("At a Glance", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Visual Summary", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Compare and Contrast", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Concept Map", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Quick FAQ", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Quick Check", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Curriculum Map", [block["title"] for block in term_payload["blocks"]])
            self.assertIn("Structure Expansion", [block["title"] for block in act_r_term["blocks"]])
            self.assertIn("Structure Expansion", [block["title"] for block in term_payload["blocks"]])
            self.assertEqual(
                next(block for block in term_payload["blocks"] if block["id"] == "study-prompts")["type"],
                "steps",
            )
            self.assertEqual(
                next(block for block in term_payload["blocks"] if block["id"] == "at-a-glance")["type"],
                "table",
            )
            self.assertEqual(
                next(block for block in term_payload["blocks"] if block["id"] == "visual-summary")["type"],
                "callout",
            )
            curriculum_block = next(block for block in term_payload["blocks"] if block["id"] == "curriculum-map")
            self.assertEqual(curriculum_block["type"], "curriculum-map")
            self.assertGreaterEqual(len(curriculum_block["sections"]), 1)
            expansion_block = next(block for block in act_r_term["blocks"] if block["id"] == "structure-expansion")
            self.assertEqual(expansion_block["type"], "structure-expansion")
            self.assertGreaterEqual(len(expansion_block["sections"]), 1)
            self.assertEqual(expansion_block["sections"][0]["layer"], "editorial-expansion")
            comparison_block = next(block for block in term_payload["blocks"] if block["id"] == "comparison")
            self.assertEqual(comparison_block["type"], "comparison")
            self.assertIn("panels", comparison_block)
            self.assertGreaterEqual(len(comparison_block["panels"]), 4)
            self.assertIn("boundary stays visible", comparison_block["panels"][2]["body"])
            diagram_block = next(block for block in term_payload["blocks"] if block["id"] == "concept-map")
            self.assertEqual(diagram_block["type"], "diagram")
            self.assertIn("center", diagram_block)
            self.assertEqual(len(diagram_block["lanes"]), 3)
            self.assertIn("takeaway", diagram_block)
            faq_block = next(block for block in term_payload["blocks"] if block["id"] == "quick-faq")
            self.assertEqual(faq_block["type"], "faq")
            self.assertGreaterEqual(len(faq_block["items"]), 3)
            self.assertIn("question", faq_block["items"][0])
            self.assertIn("answer", faq_block["items"][0])
            quiz_block = next(block for block in term_payload["blocks"] if block["id"] == "quick-quiz")
            self.assertEqual(quiz_block["type"], "quiz")
            self.assertGreaterEqual(len(quiz_block["options"]), 3)
            self.assertGreaterEqual(quiz_block["answerIndex"], 0)
            self.assertEqual(len(quiz_block["options"]), 4)
            at_a_glance = next(block for block in term_payload["blocks"] if block["id"] == "at-a-glance")
            self.assertIn("layers, activations, gradients", at_a_glance["rows"][4]["value"])
            self.assertEqual(vision_transformer_summary["taxonomy"]["topic"], "Vision Transformer")
            self.assertEqual(vision_transformer_summary["metadata"]["studyFamily"], "Computer Vision")
            self.assertEqual(near_duplicate_summary["metadata"]["studyFamily"], "Similarity & Deduplication")
            self.assertIn("Vision Transformer", transformer_summary["links"]["next"])
            self.assertIn("Transformer", vision_transformer_summary["links"]["prerequisites"])
            self.assertIn("RAG", rag_summary["aliases"])
            self.assertIn("Retrieval Augmented Generation", rag_summary["aliases"])
            self.assertEqual(paths_payload[0]["slug"], "neural-networks-advanced-architectures")
            self.assertEqual(path_payload["steps"][0]["stage"], "start")
            self.assertEqual(path_payload["steps"][0]["slug"], "transformer")

    def test_live_artifacts_match_source_workbooks(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            glossary_path = REPO_ROOT / "data_glossary.xlsx"
            structure_path = REPO_ROOT / "data_structure.xlsx"
            out_dir = Path(tmp_dir) / "published"

            subprocess.run(
                [
                    str(REPO_ROOT / ".venv" / "bin" / "python3"),
                    str(SCRIPT_PATH),
                    "--glossary-workbook",
                    str(glossary_path),
                    "--structure-workbook",
                    str(structure_path),
                    "--out-dir",
                    str(out_dir),
                ],
                check=True,
            )

            committed_report = json.loads(
                (REPO_ROOT / "public" / "content" / "published" / "reports" / "import-report.json").read_text(
                    encoding="utf-8"
                )
            )
            committed_audit = json.loads(
                (REPO_ROOT / "public" / "content" / "published" / "reports" / "content-audit.json").read_text(
                    encoding="utf-8"
                )
            )
            committed_manifest = json.loads((REPO_ROOT / "public" / "content" / "published" / "manifest.json").read_text(encoding="utf-8"))
            rebuilt_report = json.loads((out_dir / "reports" / "import-report.json").read_text(encoding="utf-8"))
            rebuilt_audit = json.loads((out_dir / "reports" / "content-audit.json").read_text(encoding="utf-8"))
            rebuilt_manifest = json.loads((out_dir / "manifest.json").read_text(encoding="utf-8"))

            committed_index = json.loads(
                (REPO_ROOT / "public" / "content" / "published" / "terms" / "index.json").read_text(
                    encoding="utf-8"
                )
            )
            rebuilt_index = json.loads((out_dir / "terms" / "index.json").read_text(encoding="utf-8"))

            committed_paths = json.loads(
                (REPO_ROOT / "public" / "content" / "published" / "paths" / "index.json").read_text(encoding="utf-8")
            )
            rebuilt_paths = json.loads((out_dir / "paths" / "index.json").read_text(encoding="utf-8"))

            self.assertEqual(len(committed_index), len(rebuilt_index))
            self.assertEqual(len(committed_paths), len(rebuilt_paths))
            self.assertEqual(committed_report["termCount"], rebuilt_report["termCount"])
            self.assertEqual(committed_report["pathCount"], rebuilt_report["pathCount"])
            self.assertEqual(committed_report["sourceInventoryTermCount"], rebuilt_report["sourceInventoryTermCount"])
            self.assertEqual(committed_audit["status"], rebuilt_audit["status"])
            self.assertEqual(committed_audit["qualityChecks"]["highSeverityIssueCount"], rebuilt_audit["qualityChecks"]["highSeverityIssueCount"])
            self.assertEqual(committed_manifest["termCount"], rebuilt_manifest["termCount"])
            self.assertEqual(committed_manifest["pathCount"], rebuilt_manifest["pathCount"])
            self.assertEqual(committed_manifest["launchSectionCount"], rebuilt_manifest["launchSectionCount"])
            self.assertEqual(committed_manifest["contentDepth"]["blockCounts"]["structure-expansion"], rebuilt_manifest["contentDepth"]["blockCounts"]["structure-expansion"])
            self.assertEqual(
                [item["slug"] for item in committed_index[:50]],
                [item["slug"] for item in rebuilt_index[:50]],
            )


def write_glossary_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "main"

    sheet["A1"] = "terms"
    sheet["D1"] = "Term from excel"
    sheet["E1"] = "vlookup"
    sheet["F1"] = "terms"

    sheet["A2"] = "Activation Function"
    sheet["A3"] = "Activation-Function"
    sheet["A4"] = "Gradient Descent"
    sheet["A5"] = "Transformer"
    sheet["A6"] = "Vision Transformer"
    sheet["A7"] = "Retrieval Augmented Generation (RAG)"
    sheet["A8"] = "Q-Transformer"
    sheet["A9"] = "Transformer Architecture"
    sheet["A10"] = "Near-duplicate detection (MinHash)"
    sheet["B10"] = "ACT R architecture"
    sheet["H3"] = "1. **Activation Function**: A non-linear transform used inside a neural network."
    sheet["N2"] = "topic"
    sheet["O2"] = "category"
    sheet["P2"] = "sub category"
    sheet["N3"] = "Activation Function"
    sheet["O3"] = "Neural Networks"
    sheet["P3"] = "Components"
    sheet["N5"] = "Transformer"
    sheet["O5"] = "Neural Networks"
    sheet["P5"] = "Advanced Architectures"
    sheet["N7"] = "Retrieval Augmented Generation (RAG)"
    sheet["O7"] = "Natural Language Processing"
    sheet["P7"] = "Knowledge-Enhanced Generation"
    sheet["N8"] = "Q-Transformer"
    sheet["O8"] = "Neural Networks"
    sheet["P8"] = "Advanced Architectures"
    sheet["N9"] = "Transformer Architecture"
    sheet["O9"] = "Neural Networks"
    sheet["P9"] = "Advanced Architectures"
    sheet["N10"] = "Near-duplicate detection (MinHash)"

    workbook.save(path)


def write_structure_workbook(path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "Sheet1"
    structure_sheet = workbook.create_sheet("Sheet2")
    structure_sheet["A2"] = ""
    structure_sheet["B2"] = "Term"
    structure_sheet["C2"] = "Introduction – Definition and Overview"
    structure_sheet["D2"] = "Applications – Real-world Use Cases and Examples"
    structure_sheet["E2"] = "Case Studies – In-depth Analysis of Real-world Applications"
    structure_sheet["A4"] = "1. Introduction"
    structure_sheet["A5"] = "2. Applications"
    workbook.save(path)


if __name__ == "__main__":
    unittest.main()
