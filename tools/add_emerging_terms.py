#!/usr/bin/env python3
"""Add 100 emerging AI/ML terms (2024-2026) to the glossary workbook."""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

GLOSSARY_PATH = Path("data_glossary.xlsx")

# 100 emerging AI/ML terms from 2024-2026 — organized by category
NEW_TERMS = [
    # ── State Space Models & New Architectures ────────────────────
    "Mamba",
    "Mamba-2",
    "Jamba",
    "RWKV",
    "StripedHyena",
    "Kolmogorov-Arnold Networks (KANs)",
    "Mixture-of-Depths",
    "Mixture-of-Recursions",
    "Linear Attention Transformer",
    "Neural State Machine",
    "Adaptive Depth Transformer",
    "Sparse Attention Mechanism",
    "Gated Linear Attention",

    # ── DPO Variants & Alignment ──────────────────────────────────
    "KTO (Kahneman-Tversky Optimization)",
    "ORPO (Odds Ratio Preference Optimization)",
    "SimPO (Simple Preference Optimization)",
    "IPO (Identity Preference Optimization)",
    "CPO (Contrastive Preference Optimization)",
    "SLiC (Sequence Likelihood Calibration)",
    "Constitutional AI",
    "Superalignment",
    "Mechanistic Interpretability",
    "Circuit Analysis",

    # ── Reasoning & Inference ─────────────────────────────────────
    "Inference-Time Compute",
    "System 2 Reasoning",
    "Test-Time Reinforcement Learning",
    "Speculative Decoding",
    "Prompt Caching",
    "Self-Correction Loop",
    "Search-Based Inference",
    "Monte Carlo Tree Search LLM",
    "Chain-of-Thought Decoding",

    # ── Frontier Models (2024-2026) ───────────────────────────────
    "GPT-4o",
    "OpenAI o1",
    "OpenAI o3",
    "Claude 3 Opus",
    "Claude 3.5 Sonnet",
    "Claude 4",
    "Gemini 1.5 Pro",
    "Gemini 1.5 Flash",
    "Gemini 2.0",
    "Llama 3",
    "Llama 3.1",
    "Llama 4",
    "DeepSeek R1",
    "DeepSeek V3",
    "Qwen2",
    "Qwen2.5",
    "Mixtral 8x22B",
    "Mistral Large",
    "Phi-3 Mini",
    "Phi-4",
    "Gemma 2",
    "Gemma 3",
    "StarCoder2",
    "Yi 34B",

    # ── Agent Frameworks & Infrastructure ─────────────────────────
    "Model Context Protocol (MCP)",
    "Microsoft Semantic Kernel",
    "GPTScript",
    "AgentOps",
    "LLMOps Pipeline",
    "AI Browser Agent",
    "Computer Use (Anthropic)",
    "Computer Use (OpenAI)",
    "Mem0 Memory Layer",
    "Zep Memory Layer",
    "Episodic Memory Module",

    # ── Multimodal & Video Generation ─────────────────────────────
    "Sora (OpenAI Video)",
    "Kling Video Generation",
    "Gen-3 Alpha (Runway)",
    "Stable Video Diffusion",
    "Chameleon (Meta)",
    "Pixtral (Mistral)",
    "Gemini Omni",
    "Audio-to-Audio Streaming",
    "Interleaved Image-Text Training",
    "Native Multimodal Embedding",

    # ── Embodied AI & Robotics ────────────────────────────────────
    "Gemini Robotics 1.5",
    "Chain-of-Action Planning",
    "Foundation Model for Robotics",
    "Molmo-Act",

    # ── Scientific Discovery ──────────────────────────────────────
    "DeepMind Co-Scientist",
    "ProGen3 Protein Generation",
    "Virtual Lab Simulation",
    "AI-Driven Drug Discovery",

    # ── AI Safety & Governance ────────────────────────────────────
    "EU AI Act Compliance",
    "Automated Red Teaming",
    "Model Transparency Budget",
    "Pragmatic Safety Budget",
    "Capability Tradeoff Analysis",

    # ── Efficiency & Optimization ─────────────────────────────────
    "FP8 Quantization",
    "INT4 Quantization",
    "EXL2 Quantization Format",
    "GGUF Quantization Format",
    "Precision-Aware Training",
    "Adaptive Quantization",
    "Token Merging",

    # ── Agentic Patterns ──────────────────────────────────────────
    "Multi-Agent Orchestration",
    "Agent Swarm",
    "Reflection Agent Pattern",
    "Tool-Use Agent Loop",
    "ReAct Agent Pattern",
    "Plan-and-Execute Agent",
    "Agentic RAG",

    # ── Data & Scaling ────────────────────────────────────────────
    "Synthetic Data Generation Loop",
    "High-Quality Curation Pipeline",
    "Sovereign AI Infrastructure",
    "Stargate Supercomputer",
    "Post-Training Scaling Laws",
]


def main() -> None:
    """Append new terms to the glossary workbook's main sheet."""
    wb = load_workbook(GLOSSARY_PATH)
    sheet = wb["main"]

    # Find last populated row
    last_row = 1
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        if row[0]:
            last_row += 1

    print(f"Current last data row: {last_row}")
    print(f"Adding {len(NEW_TERMS)} new terms...")

    added = 0
    for i, term in enumerate(NEW_TERMS):
        new_row = last_row + 1 + i
        sheet.cell(row=new_row, column=1, value=term)  # Column A
        added += 1

    wb.save(GLOSSARY_PATH)
    print(f"Saved workbook. Added {added} new terms starting at row {last_row + 1}.")


if __name__ == "__main__":
    main()
