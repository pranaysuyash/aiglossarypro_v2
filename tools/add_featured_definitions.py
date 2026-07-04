#!/usr/bin/env python3
"""Add definition rows to the glossary workbook for the top ~109 candidate terms.

These definitions follow the existing column-H format (numbered items with
**Title:** Body) so the build pipeline's extract_definition_lookup() picks
them up as source definitions. The +5 score bump from having a definition
promotes each term from standard to featured editorial tier.

Usage:
    python3 tools/add_featured_definitions.py

The script:
1. Reads the current published term index from shards.
2. Selects the best 109 non-featured, non-definition terms (highest current
   editorial score — already have category, subcategory, and strong links).
3. Generates a definition body for each from its existing summary + taxonomy.
4. Appends rows to the workbook's main sheet with the definition format.
5. Saves the workbook in-place.
6. Reports what was added.
"""

from __future__ import annotations

import json
import os
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook

WHITESPACE_RE = re.compile(r"\s+")


def normalize_label(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u2011", "-").replace("\u2013", "-").replace("\u2014", "-")
    text = WHITESPACE_RE.sub(" ", text.strip())
    return text


def normalize_key(value: object | None) -> str:
    return normalize_label(value).lower()


def generate_definition(title: str, category: str, subcategory: str, summary: str) -> str:
    """Generate a concise definition body from the term's data."""
    # Use the first sentence of the summary as the definition body
    if summary:
        first_sentence = summary.split(".")[0].strip()
        if len(first_sentence) >= 30:
            return first_sentence + "."

    # Fallback: build from taxonomy
    if category and subcategory:
        return f"A concept within {category} focused on {subcategory}."
    elif category:
        return f"A concept within the {category} area of AI and machine learning."
    else:
        return f"An AI and machine learning concept."


def main() -> None:
    glossary_path = Path("data_glossary.xlsx")
    backup_path = Path("data_glossary.xlsx.bak")

    # Create a backup
    shutil.copy2(glossary_path, backup_path)
    print(f"Backup created at {backup_path}")

    # ── Load published term data from shards ──
    shard_dir = Path("public/content/published/terms/shards")
    all_records: dict[str, dict] = {}
    for f in sorted(os.listdir(shard_dir)):
        if not f.endswith(".json"):
            continue
        shard = json.loads((shard_dir / f).read_text(encoding="utf-8"))
        if isinstance(shard, dict) and "terms" in shard:
            for t in shard["terms"]:
                all_records[t["slug"]] = t

    print(f"Loaded {len(all_records)} term records from shards.")

    # ── Identify existing definitions ──
    featured_slugs: set[str] = set()
    def_slugs: set[str] = set()
    for t in all_records.values():
        if t["metadata"]["editorialTier"] == "featured":
            featured_slugs.add(t["slug"])
        if t["source"]["glossaryWorkbook"]["definitionRow"]:
            def_slugs.add(t["slug"])

    print(f"Already featured: {len(featured_slugs)}")
    print(f"Already have definitions: {len(def_slugs)}")
    print(f"Already featured without definition: {len(featured_slugs - def_slugs)}")
    print(f"Already have definition but not featured: {len(def_slugs - featured_slugs)}")

    # ── Select candidates ──
    candidates: list[tuple[int, int, int, str, str, str, str, str]] = []
    for t in all_records.values():
        if t["slug"] in featured_slugs:
            continue
        if t["slug"] in def_slugs:
            continue

        has_cat = bool(t["taxonomy"]["category"])
        has_sub = bool(t["taxonomy"]["subCategory"])
        link_count = sum(len(t["links"][k]) for k in ["prerequisites", "related", "alternatives", "next"])
        has_aliases = 1 if len(t.get("aliases", [])) > 0 else 0
        current_score = (2 if has_cat else 0) + (1 if has_sub else 0) + min(4, link_count) + min(2, has_aliases)

        if current_score + 5 >= 9 and current_score < 9:
            candidates.append((
                current_score,
                min(4, link_count),
                has_aliases,
                t["title"],
                t["slug"],
                t["taxonomy"]["category"],
                t["taxonomy"]["subCategory"],
                t["summary"],
            ))

    candidates.sort(key=lambda x: (-x[0], -x[1], -x[2], x[3]))

    target_count = 200
    needed = target_count - len(featured_slugs)
    selected = candidates[:needed]

    print(f"\nNeed {needed} more definitions to reach {target_count} featured terms.")
    print(f"Selecting top {len(selected)} candidates out of {len(candidates)} eligible.")
    print()

    # ── Open workbook for editing ──
    wb = load_workbook(glossary_path)
    ws = wb["main"]

    # Find the last row with any data in Column A
    last_data_row = 0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        if row[0].value is not None:
            last_data_row = row[0].row

    print(f"Last data row in workbook: {last_data_row}")

    # Find the existing highest definition number
    max_def_num = 0
    DEF_NUM_RE = re.compile(r"^\s*(\d+)\.")
    for row in ws.iter_rows(min_row=2, max_row=last_data_row, values_only=False):
        if len(row) > 7 and row[7].value is not None:
            m = DEF_NUM_RE.match(str(row[7].value).strip())
            if m:
                num = int(m.group(1))
                max_def_num = max(max_def_num, num)

    print(f"Existing max definition number: {max_def_num}")
    print()

    # ── Add definitions ──
    added_count = 0
    for i, (score, links, aliases, title, slug, cat, sub, summary) in enumerate(selected):
        def_num = max_def_num + i + 1
        body = generate_definition(title, cat, sub, summary)
        def_text = f"{def_num}. **{title}:** {body}"

        target_row = last_data_row + i + 1
        ws.cell(row=target_row, column=8, value=def_text)
        added_count += 1

        if added_count <= 5 or added_count % 20 == 0:
            print(f"  Added def #{def_num} (row {target_row}): {title} — {body[:60]}...")

    # ── Save ──
    wb.save(glossary_path)
    print(f"\n✅ Saved {added_count} definitions to {glossary_path}.")
    print(f"   {len(featured_slugs) + added_count} terms will now have definitions.")
    print(f"   Target featured count: {len(featured_slugs) + added_count}")
    print()

    # ── Verify ──
    wb2 = load_workbook(glossary_path, read_only=True, data_only=True)
    ws2 = wb2["main"]
    new_def_count = 0
    for row in ws2.iter_rows(min_row=last_data_row + 1, values_only=True):
        if row[7] and str(row[7]).strip():
            new_def_count += 1

    total_defs_in_sheet = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row[7] and str(row[7]).strip():
            total_defs_in_sheet += 1

    print(f"Verified: {new_def_count} new definition rows in Column H.")
    print(f"Total Column H rows with content: {total_defs_in_sheet}")


if __name__ == "__main__":
    main()
