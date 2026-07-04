#!/usr/bin/env python3
"""Build published glossary artifacts from the real source workbooks.

This importer is intentionally workbook-aware rather than assuming a clean
row-aligned table:

- glossary workbook `main` columns A, B, and D:G provide the broad term inventory
- glossary workbook `main` columns N:P are a smaller taxonomy lookup keyed by term
- glossary workbook `main` column H contains a small set of inline definition rows
- structure workbook `Sheet2` row 2 contains the editorial field registry

The goal is to publish the real source corpus into the runtime JSON shape while
preserving source evidence and coverage gaps honestly.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import tempfile
from datetime import datetime, timezone
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


WHITESPACE_RE = re.compile(r"\s+")
NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")
TOKEN_RE = re.compile(r"[a-z0-9]+")
DEFINITION_NUMBERED_RE = re.compile(
    r"^\s*\d+\.\s+(?:\*\*(?P<numbered_title>.+?)\*\*|(?P<numbered_title_plain>[^:]+)):\s*(?P<body>.+?)\s*$"
)
DEFINITION_FALLBACK_RE = re.compile(r"^\s*[*]{0,2}(?P<title>.*?)\*{0,2}\s*:\s*(?P<body>.+?)\s*$")
DEFINITION_BODY_MIN_CHARS = 30
DEFINITION_TITLE_MAX_CHARS = 160
EXPECTED_BLOCK_SEQUENCE = ("overview", "taxonomy", "connections", "study-prompts")
INVENTORY_COLUMNS = (1, 2, 4, 5, 6, 7)
INVENTORY_REJECT_KEYS = {"terms", "term", "topic", "n/a", "#n/a", "n-a", "na"}
STRUCTURE_LAUNCH_SECTION_ORDER = (
    "Introduction",
    "Prerequisites",
    "Theoretical Concepts",
    "How It Works",
    "Variants or Extensions",
    "Applications",
    "Implementation",
    "Evaluation and Metrics",
    "Advantages and Disadvantages",
    "Ethics and Responsible AI",
    "Related Concepts",
)
STRUCTURE_LAUNCH_SECTIONS = set(STRUCTURE_LAUNCH_SECTION_ORDER)
STRUCTURE_EDITORIAL_SECTION_ORDER = (
    "Historical Context",
    "Case Studies",
    "Hands-on Tutorials",
    "Industry Insights",
    "Common Challenges and Pitfalls",
    "Real-world Datasets and Benchmarks",
    "Tools and Frameworks",
    "Further Reading",
    "Research Papers",
    "Career Guidance",
    "Future Directions",
    "Glossary",
    "FAQs",
    "Tags and Keywords",
    "References",
    "Conclusion",
    "Best Practices",
    "Security Considerations",
    "Optimization Techniques",
    "Comparison with Alternatives",
    "Illustration or Diagram",
)
STRUCTURE_EDITORIAL_SECTIONS = set(STRUCTURE_EDITORIAL_SECTION_ORDER)
STRUCTURE_LAYER_PRIORITY = {
    "launch-runtime": 0,
    "editorial-expansion": 1,
    "backlog": 2,
}
FAMILY_CONTENT_PROFILES = {
    "Computer Vision": {
        "note": "Watch how the concept changes image, video, spatial, or scene understanding instead of treating it as a generic AI label.",
        "boundary": "Ask whether the term changes pixel-level, region-level, object-level, or scene-level interpretation.",
        "quiz_anchor": "Use the visual pipeline to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with nearby vision methods so the boundary stays visible across image, video, and scene tasks.",
    },
    "Natural Language Processing": {
        "note": "Watch how the concept changes token flow, retrieval, generation, or evaluation rather than the surface wording alone.",
        "boundary": "Ask whether the term changes tokenization, representation, retrieval, generation, or evaluation behavior.",
        "quiz_anchor": "Use the language pipeline to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with nearby NLP methods so you can separate retrieval, generation, and evaluation roles.",
    },
    "Reinforcement Learning": {
        "note": "Watch how the concept changes policy choice, reward structure, exploration, or control behavior.",
        "boundary": "Ask whether the term changes policy, reward, exploration, or control dynamics.",
        "quiz_anchor": "Use the agent loop to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with adjacent RL ideas so the policy and reward boundary stays sharp.",
    },
    "Statistics": {
        "note": "Watch the assumptions, uncertainty, and inference boundaries before memorizing a formal definition.",
        "boundary": "Ask which assumptions, uncertainty sources, or inference boundaries the term is really changing.",
        "quiz_anchor": "Use the inference setup to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with nearby statistical ideas so you can separate assumptions from conclusions.",
    },
    "Evaluation": {
        "note": "Watch what is measured, which baseline is implied, and what tradeoff the metric is trying to improve.",
        "boundary": "Ask what is measured, what is not measured, and which tradeoff the metric rewards.",
        "quiz_anchor": "Use the evaluation setup to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with adjacent metrics so the baseline and tradeoff stay visible.",
    },
    "Ethics & Governance": {
        "note": "Watch the safety, fairness, privacy, and accountability implications before treating it as a neutral term.",
        "boundary": "Ask which risk, harm, or accountability concern the term is meant to surface.",
        "quiz_anchor": "Use the safety and governance lens to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with adjacent governance ideas so the responsibility boundary stays clear.",
    },
    "Similarity & Deduplication": {
        "note": "Watch the normalization, match threshold, and false-positive or false-negative balance.",
        "boundary": "Ask which normalization, threshold, or matching decision changes the result.",
        "quiz_anchor": "Use the dedupe pipeline to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with nearby matching methods so threshold and normalization differences stay visible.",
    },
    "Neural Networks": {
        "note": "Watch how the concept affects layers, activations, gradients, optimization, or training dynamics.",
        "boundary": "Ask whether the term changes layers, activations, gradients, optimization, or training dynamics.",
        "quiz_anchor": "Use the network pipeline to situate it before you memorize the wording.",
        "compare_anchor": "Compare it with adjacent neural-network ideas so the layer and gradient boundary stays visible.",
    },
}


# ── Auto-classification system ──────────────────────────────────────────────

TAXONOMY_COVERAGE_MINIMUM = 0.74
PUBLISHED_SCHEMA_VERSION = "published-manifest.v1"
CONTENT_VERSION = "2026.07.04-source-v1"
BUILDER_VERSION = "build_published_content.atomic-v1"

AUTO_CLASSIFICATION_RULES: list[tuple[set[str], str, str, int]] = [
    # ==========================================
    # ORIGINAL RULES (curated, high confidence)
    # ==========================================
    ({"weight", "parameter"}, "Neural Networks", "Model Parameters", 8),
    ({"ai learning"}, "AI Applications", "", 10),
    ({"ai generated"}, "AI Applications", "", 10),
    ({"few shot"}, "Learning Paradigms", "Low-Data Learning", 10),
    ({"zero shot"}, "Learning Paradigms", "Low-Data Learning", 10),
    ({"one shot"}, "Learning Paradigms", "Low-Data Learning", 10),
    ({"chain of thought", "chain of thought"}, "Natural Language Processing", "Language Generation", 10),
    ({"actor critic"}, "Reinforcement Learning", "Advanced RL Techniques", 10),
    ({"retrieval augmented"}, "Natural Language Processing", "Pre-trained Models", 10),
    ({"text to image"}, "Generative Models", "Diffusion Models", 10),
    ({"text to speech"}, "Natural Language Processing", "Language Generation", 10),
    ({"speech to text"}, "Natural Language Processing", "Language Generation", 10),
    ({"image to image"}, "Computer Vision", "Image Processing", 10),
    ({"adversarial attack"}, "Security & Robustness", "Adversarial Machine Learning", 10),
    ({"adversarial example"}, "Security & Robustness", "Adversarial Machine Learning", 10),
    ({"federated learning"}, "Learning Paradigms", "Low-Data Learning", 10),
    ({"differential privacy", "privacy preserving"}, "Ethics & Governance", "Fairness in ML", 9),
    ({"explainable ai", "xai"}, "Explainable AI", "Model Interpretability", 10),
    ({"time series", "time series"}, "Data Processing", "Data Pipeline", 9),
    ({"reinforcement learning"}, "Reinforcement Learning", "Learning Techniques", 10),
    ({"semi supervised"}, "Learning Paradigms", "Low-Data Learning", 10),
    ({"self supervised"}, "Learning Paradigms", "Low-Data Learning", 10),
    ({"active learning"}, "Learning Paradigms", "Low-Data Learning", 9),
    ({"multi task", "multitask"}, "Neural Networks", "Training Techniques", 9),
    ({"multi modal", "multimodal"}, "Neural Networks", "Specialized Architectures", 9),
    ({"out of distribution", "ood"}, "Model Evaluation", "Generalization Issues", 9),
    ({"bag of words"}, "Representation Learning", "Vector Representations", 9),
    ({"word embedding"}, "Representation Learning", "Vector Representations", 9),
    ({"dimensionality reduction"}, "Dimensionality Reduction", "Feature Reduction", 10),
    ({"representation learning"}, "Representation Learning", "Feature Learning", 10),
    ({"feature extraction"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"feature engineering"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"feature selection"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"hyperparameter tuning", "hyperparameter optimization"}, "Optimization Algorithms", "Gradient-based Optimizers", 9),
    ({"model compression"}, "Neural Networks", "Training Techniques", 9),
    ({"knowledge distillation"}, "Neural Networks", "Training Techniques", 9),
    ({"model quantization"}, "Neural Networks", "Training Techniques", 9),
    ({"ablation study"}, "Model Evaluation", "Performance Metrics", 9),
    ({"anomaly detection"}, "Detection Techniques", "Unsupervised Detection", 9),
    ({"object detection"}, "Computer Vision", "Object Detection", 9),
    ({"image segmentation"}, "Computer Vision", "Image Processing", 9),
    ({"image classification"}, "Computer Vision", "Image Processing", 9),
    ({"sentiment analysis"}, "Natural Language Processing", "Text Analysis", 9),
    ({"named entity", "named entity recognition"}, "Natural Language Processing", "Text Analysis", 9),
    ({"text classification"}, "Natural Language Processing", "Text Analysis", 9),
    ({"machine translation"}, "Natural Language Processing", "Language Generation", 9),
    ({"question answering"}, "Natural Language Processing", "Language Generation", 9),
    ({"text summarization"}, "Natural Language Processing", "Language Generation", 9),
    ({"image captioning"}, "Computer Vision", "Image Processing", 8),
    ({"style transfer"}, "Computer Vision", "Image Processing", 8),
    ({"super resolution"}, "Computer Vision", "Image Processing", 8),
    ({"image inpainting"}, "Computer Vision", "Image Processing", 8),
    ({"image generation"}, "Generative Models", "Adversarial Networks", 9),
    ({"text generation"}, "Natural Language Processing", "Language Generation", 9),
    ({"code generation"}, "Natural Language Processing", "Language Generation", 9),
    ({"data augmentation"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"data cleaning"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"data labeling"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"cross validation"}, "Model Evaluation", "Performance Metrics", 9),
    ({"transfer learning"}, "Transfer Learning", "Model Adaptation", 10),
    ({"fine tuning", "finetune", "fine tuning", "finetuning"}, "Transfer Learning", "Model Adaptation", 10),
    ({"domain adaptation"}, "Transfer Learning", "Model Adaptation", 9),
    ({"ensemble learning"}, "Learning Paradigms", "Supervised Learning", 9),
    ({"deep learning"}, "Neural Networks", "Core Models", 9),
    ({"machine learning"}, "Foundations", "Core Concepts", 9),
    ({"neural network"}, "Neural Networks", "Core Models", 9),
    ({"convolutional", "convolution"}, "Neural Networks", "Core Models", 8),
    ({"recurrent", "rnn"}, "Neural Networks", "Recurrent Models", 8),
    ({"long short term", "lstm"}, "Neural Networks", "Recurrent Models", 8),
    ({"gated recurrent", "gru"}, "Neural Networks", "Recurrent Models", 8),
    ({"transformer"}, "Neural Networks", "Core Models", 8),
    ({"attention mechanism"}, "Neural Networks", "Attention Mechanism", 9),
    ({"self attention"}, "Neural Networks", "Attention Mechanism", 9),
    ({"generative adversarial", "gan"}, "Generative Models", "Adversarial Networks", 9),
    ({"variational autoencoder", "vae"}, "Generative Models", "Autoencoders", 9),
    ({"diffusion model"}, "Generative Models", "Diffusion Models", 9),
    ({"autoencoder"}, "Neural Networks", "Autoencoders", 8),
    ({"gradient descent"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"stochastic gradient", "sgd"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"adam optimizer", "adamw", "adam"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"batch normalization"}, "Neural Networks", "Training Techniques", 8),
    ({"layer normalization"}, "Neural Networks", "Training Techniques", 8),
    ({"dropout"}, "Neural Networks", "Regularization Techniques", 8),
    ({"regularization"}, "Neural Networks", "Regularization Techniques", 8),
    ({"backpropagation", "backprop"}, "Neural Networks", "Training Algorithms", 8),
    ({"loss function"}, "Loss Functions", "Classification Loss", 8),
    ({"activation function"}, "Neural Networks", "Activation Functions", 8),
    ({"metric learning"}, "Representation Learning", "Feature Learning", 8),
    ({"contrastive learning"}, "Representation Learning", "Feature Learning", 8),
    ({"recommender system"}, "AI Applications", "Knowledge Representation", 8),
    ({"knowledge graph"}, "AI Theory", "Knowledge Representation", 8),
    ({"graph neural", "gnn"}, "Neural Networks", "Graph-based Models", 8),
    ({"graph convolution"}, "Neural Networks", "Graph-based Models", 8),
    ({"graph attention"}, "Neural Networks", "Graph-based Models", 8),
    ({"natural language", "tokenizer", "tokenization", "lemmatization", "stemming"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"word2vec", "glove", "fasttext", "cbow", "skip gram"}, "Representation Learning", "Vector Representations", 8),
    ({"bert", "gpt", "t5", "bart", "roberta", "electra", "xlnet", "albert"}, "Natural Language Processing", "Pre-trained Models", 8),
    ({"resnet", "densenet", "mobilenet", "efficientnet", "inception", "vgg", "yolo", "rcnn", "unet", "detr", "segformer"}, "Computer Vision", "Core Models", 8),
    ({"vit", "vision transformer"}, "Computer Vision", "Core Models", 8),
    ({"clip", "contrastive language image pretraining"}, "Computer Vision", "Core Models", 8),
    ({"reinforce", "policy gradient", "q learning", "deep q", "ppo", "sac", "td3"}, "Reinforcement Learning", "Advanced RL Techniques", 8),
    ({"monte carlo"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"markov chain", "markov decision"}, "AI Theory", "Knowledge Representation", 8),
    ({"bayesian network", "probabilistic graphical"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"gaussian process"}, "Statistical Models", "Probabilistic Models", 8),
    ({"random forest"}, "Learning Paradigms", "Supervised Learning", 8),
    ({"decision tree"}, "Learning Paradigms", "Supervised Learning", 8),
    ({"support vector", "svm"}, "Learning Paradigms", "Supervised Learning", 8),
    ({"k nearest", "knn"}, "Learning Paradigms", "Supervised Learning", 8),
    ({"k means", "kmeans"}, "Detection Techniques", "Unsupervised Detection", 8),
    ({"pca", "principal component"}, "Dimensionality Reduction", "Feature Reduction", 8),
    ({"tsne", "t sne"}, "Dimensionality Reduction", "Feature Reduction", 8),
    ({"gradient boosting", "xgboost", "lightgbm", "catboost"}, "Machine Learning Frameworks", "Automated Processes", 8),
    ({"database", "sql", "nosql", "vector database"}, "Data Processing", "Data Pipeline", 7),
    ({"api", "microservice", "rest", "graphql", "endpoint"}, "Data Processing", "Data Pipeline", 7),
    ({"gpu", "tpu", "cuda", "tensor core"}, "Neural Networks", "Training Techniques", 7),
    ({"cloud", "aws", "gcp", "azure", "kubernetes", "docker"}, "AI Applications", "", 7),
    ({"edge ai", "edge computing", "on device"}, "AI Applications", "", 7),
    ({"robotics", "robot", "autonomous", "drone"}, "AI Applications", "", 8),
    ({"healthcare", "medical", "clinical", "diagnosis", "drug"}, "AI Applications", "", 8),
    ({"finance", "financial", "trading", "portfolio"}, "AI Applications", "", 8),
    ({"cybersecurity", "cyber"}, "AI Applications", "", 8),
    ({"nlp", "llm", "ner", "nlu", "asr", "tts"}, "Natural Language Processing", "", 8),
    ({"cnn", "vision", "image", "video", "pixel"}, "Computer Vision", "", 8),

    # ==========================================
    # LLM-DERIVED RULES — AI Applications broad patterns
    # ==========================================
    ({"ai in", "ai for", "ai and", "ai on"}, "AI Applications", "", 9),
    ({"ai research", "ai ethics", "ai safety", "ai alignment", "ai governance"}, "AI Applications", "", 8),
    ({"ai system", "ai systems", "ai model", "ai models", "ai tool", "ai tools", "ai platform"}, "AI Applications", "", 8),
    ({"ai driven", "ai based", "ai powered"}, "AI Applications", "", 8),

    # === Cognitive / Agent Systems ===
    ({"cognitive", "cognition", "act r", "soar", "bdi"}, "AI Theory", "Knowledge Representation", 9),
    ({"agent", "agentic", "multi agent", "autonomous agent", "intelligent agent"}, "AI Applications", "", 9),
    ({"autogen", "crewai", "langgraph", "pydantic ai"}, "AI Applications", "", 8),
    ({"auto gpt", "autogpt", "babyagi", "gpt researcher"}, "AI Applications", "", 8),
    ({"alphafold", "alphago", "alphazero", "alphastar", "openai five"}, "AI Applications", "", 8),
    ({"claude", "gpt 4", "gpt4", "gpt 3", "gpt3", "gemini", "mistral", "llama"}, "Natural Language Processing", "Pre-trained Models", 8),
    ({"copilot", "codex"}, "Natural Language Processing", "Language Generation", 8),

    # === Domain applications ===
    ({"education", "teaching", "learning tool", "tutor"}, "AI Applications", "", 8),
    ({"legal", "law", "compliance"}, "AI Applications", "", 8),
    ({"climate", "environmental", "sustainability", "energy"}, "AI Applications", "", 8),
    ({"gaming", "game ai", "procedural generation", "npc"}, "AI Applications", "", 8),
    ({"autonomous vehicle", "self driving", "driverless"}, "AI Types", "Narrow AI", 8),

    # === Time series / signal ===
    ({"time series", "temporal", "sequential data"}, "Data Processing", "Data Pipeline", 8),

    # === Techniques ===
    ({"ablation", "sensitivity analysis"}, "Model Evaluation", "Performance Metrics", 8),
    ({"affine transformation", "homography"}, "Computer Vision", "Image Processing", 8),
    ({"auto differentiation", "autograd", "autodiff"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"beam search", "greedy decoding"}, "Natural Language Processing", "Language Generation", 8),
    ({"bilinear", "trilinear", "interpolation"}, "Computer Vision", "Image Processing", 8),
    ({"focal loss", "dice loss", "huber loss", "hinge loss"}, "Loss Functions", "Classification Loss", 8),
    ({"multi head", "multihead"}, "Neural Networks", "Attention Mechanism", 8),
    ({"positional encoding"}, "Neural Networks", "Core Models", 8),
    ({"residual connection", "skip connection", "residual block"}, "Neural Networks", "Core Models", 8),
    ({"softmax", "log softmax"}, "Neural Networks", "Activation Functions", 8),
    ({"teacher forcing"}, "Natural Language Processing", "Language Generation", 8),
    ({"top k", "top p", "temperature scaling"}, "Natural Language Processing", "Language Generation", 8),
    ({"label smoothing"}, "Neural Networks", "Regularization Techniques", 8),
    ({"early stopping"}, "Neural Networks", "Training Techniques", 8),
    ({"learning rate", "lr scheduler", "lr schedule"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"weight decay", "weight tying"}, "Neural Networks", "Model Parameters", 8),
    ({"gradient clipping"}, "Neural Networks", "Training Techniques", 8),
    ({"data parallel", "model parallel", "distributed training"}, "Neural Networks", "Training Techniques", 8),
    ({"mixed precision", "fp16", "bf16", "amp"}, "Neural Networks", "Training Techniques", 8),
    ({"curriculum learning"}, "Learning Paradigms", "Supervised Learning", 8),
    ({"imitation learning", "behavior cloning", "behavioral cloning"}, "Reinforcement Learning", "Learning Techniques", 8),
    ({"inverse reinforcement"}, "Reinforcement Learning", "Advanced RL Techniques", 8),
    ({"hierarchical reinforcement", "hierarchical rl"}, "Reinforcement Learning", "Advanced RL Techniques", 8),

    # === Affective / social AI ===
    ({"affective computing", "emotion ai", "emotion recognition", "sentiment"}, "AI Applications", "", 8),

    # === Social / behavioral AI ===
    ({"social network", "community detection", "influence"}, "AI Applications", "", 8),

    # === Deception / red-team ===
    ({"red team", "jailbreak", "prompt injection"}, "Security & Robustness", "Adversarial Machine Learning", 9),

    # === Meta-learning ===
    ({"meta learning", "learn to learn"}, "Learning Paradigms", "Low-Data Learning", 9),

    # === Alignment / safety ===
    ({"superalignment", "mechanistic interpretability", "circuit analysis"}, "Explainable AI", "Model Interpretability", 9),

    # === Fine-tuning / adaptation patterns ===
    ({"lora", "qlora", "adapter", "prefix tuning", "prompt tuning", "p tuning"}, "Transfer Learning", "Model Adaptation", 9),

    # === Specific named techniques ===
    ({"chain of thought", "tree of thought", "reasoning"}, "Natural Language Processing", "Language Generation", 9),
    ({"rag", "retrieval augmented generation"}, "Natural Language Processing", "Pre-trained Models", 9),
    ({"instruction tuning", "instruction following"}, "Natural Language Processing", "Pre-trained Models", 8),
    ({"rlhf", "reinforcement learning from human feedback"}, "Natural Language Processing", "Pre-trained Models", 9),
    ({"dpo", "direct preference optimization"}, "Natural Language Processing", "Pre-trained Models", 9),

    # === AI safety / bias ===
    ({"hallucination", "confabulation"}, "Natural Language Processing", "Language Generation", 8),

    # ==========================================
    # LLM-DERIVED RULES Batch 2 — Broad single-token patterns
    # These tokens appear in hundreds of unclassified terms.
    # Priority 9 for high-confidence, 8 for medium-confidence.
    # ==========================================

    # === Neural Networks / Architecture ===
    ({"neural"}, "Neural Networks", "Core Models", 9),
    ({"graph"}, "Neural Networks", "Graph-based Models", 9),
    ({"networks"}, "Neural Networks", "Core Models", 8),
    ({"transformers"}, "Neural Networks", "Core Models", 8),
    ({"attention"}, "Neural Networks", "Attention Mechanism", 8),
    ({"gradient"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"differentiable"}, "Neural Networks", "Training Techniques", 9),
    ({"scaling"}, "Neural Networks", "Training Techniques", 8),
    ({"pruning"}, "Neural Networks", "Training Techniques", 8),
    ({"sparse"}, "Neural Networks", "Model Parameters", 8),
    ({"memory"}, "Neural Networks", "Specialized Architectures", 8),

    # === Data / Features ===
    ({"feature"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"embedding"}, "Representation Learning", "Vector Representations", 8),
    ({"latent"}, "Representation Learning", "Feature Learning", 8),

    # === Statistical Methods ===
    ({"bayesian"}, "Statistical Methods", "Bayesian Inference", 9),
    ({"regression"}, "Statistical Models", "Probabilistic Models", 9),
    ({"probabilistic"}, "Statistical Models", "Probabilistic Models", 9),
    ({"clustering"}, "Detection Techniques", "Unsupervised Detection", 9),
    ({"causal"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"inference"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"bootstrap"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"similarity"}, "Similarity & Deduplication", "", 8),
    ({"sampling"}, "Statistical Methods", "Bayesian Inference", 8),

    # === Security & Robustness ===
    ({"adversarial"}, "Security & Robustness", "Adversarial Machine Learning", 9),

    # === NLP / Language ===
    ({"language"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"parsing"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"semantic"}, "Natural Language Processing", "Text Analysis", 8),
    ({"coreference"}, "Natural Language Processing", "Text Analysis", 9),
    ({"slot"}, "Natural Language Processing", "Language Generation", 8),

    # === Learning Paradigms ===
    ({"ensemble"}, "Learning Paradigms", "Supervised Learning", 8),
    ({"trees"}, "Learning Paradigms", "Supervised Learning", 8),

    # === AI Applications ===
    ({"quantum"}, "AI Applications", "", 8),
    ({"heuristic"}, "AI Theory", "Knowledge Representation", 8),
    ({"planning"}, "AI Applications", "", 8),
    ({"agents"}, "AI Applications", "", 8),
    ({"temporal"}, "Data Processing", "Data Pipeline", 8),

    # ==========================================
    # LLM-DERIVED RULES Batch 3 — X-pattern multi-word patterns
    # Catch terms with suffixes like "X-based", "X-aware", "cross-X",
    # "X-driven", "X-free", "X-wise", "X-level" etc.
    # With hyphen normalization, "attention based" now matches
    # "Attention-Based" in titles. Using space-separated multi-word
    # keywords for readability.
    # ==========================================

    # === X-based (142 unique unclassified terms) ===
    ({"attention based"}, "Neural Networks", "Attention Mechanism", 9),
    ({"transformer based"}, "Neural Networks", "Core Models", 9),
    ({"graph based"}, "Neural Networks", "Graph-based Models", 9),
    ({"memory based"}, "Neural Networks", "Specialized Architectures", 8),
    ({"rule based"}, "Foundations", "Core Concepts", 8),
    ({"embedding based"}, "Representation Learning", "Vector Representations", 8),
    ({"image based"}, "Computer Vision", "", 8),
    ({"text based"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"token based"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"agent based"}, "AI Applications", "", 8),
    ({"model based"}, "AI Applications", "", 8),
    ({"data based"}, "Data Preprocessing", "Data Enhancement", 8),
    ({"kernel based"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"similarity based"}, "Similarity & Deduplication", "", 8),
    ({"feature based"}, "Data Preprocessing", "Data Enhancement", 8),
    ({"score based"}, "Generative Models", "Diffusion Models", 8),
    ({"energy based"}, "Generative Models", "", 8),
    ({"flow based"}, "Generative Models", "", 8),
    ({"rl based"}, "Reinforcement Learning", "Learning Techniques", 8),
    ({"nn based"}, "Neural Networks", "Core Models", 8),
    ({"cnn based"}, "Computer Vision", "Core Models", 8),
    ({"lstm based"}, "Neural Networks", "Recurrent Models", 8),

    # === X-aware (56 unique unclassified terms) ===
    ({"context aware"}, "Natural Language Processing", "Text Analysis", 8),
    ({"privacy aware"}, "Ethics & Governance", "Fairness in ML", 8),
    ({"carbon aware"}, "AI Applications", "", 8),
    ({"confidence aware"}, "Model Evaluation", "Generalization Issues", 8),
    ({"domain aware"}, "Transfer Learning", "Domain Adaptation", 8),
    ({"task aware"}, "Transfer Learning", "Model Adaptation", 8),
    ({"cost aware"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"uncertainty aware"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"risk aware"}, "AI Applications", "", 8),
    ({"safety aware"}, "Ethics & Governance", "", 8),
    ({"bias aware"}, "Ethics & Governance", "Fairness in ML", 8),
    ({"communication aware"}, "AI Applications", "", 8),

    # === cross-X (37 unique unclassified terms) ===
    ({"cross entropy"}, "Loss Functions", "Classification Loss", 9),
    ({"cross modal"}, "Neural Networks", "Specialized Architectures", 8),
    ({"cross domain"}, "Transfer Learning", "Domain Adaptation", 8),
    ({"cross lingual"}, "Natural Language Processing", "Text Analysis", 8),
    ({"cross task"}, "Transfer Learning", "Model Adaptation", 8),
    ({"cross attention"}, "Neural Networks", "Attention Mechanism", 8),
    ({"cross platform"}, "AI Applications", "", 7),

    # === X-driven (10 unique unclassified terms) ===
    ({"data driven"}, "Data Preprocessing", "Data Enhancement", 8),
    ({"curiosity driven"}, "Reinforcement Learning", "Learning Techniques", 8),
    ({"goal driven"}, "AI Applications", "", 8),
    ({"evidence driven"}, "AI Applications", "", 7),

    # === X-free (26 unique unclassified terms) ===
    ({"label free"}, "Data Preprocessing", "Data Enhancement", 8),
    ({"annotation free"}, "Data Preprocessing", "Data Enhancement", 8),
    ({"anchor free"}, "Computer Vision", "Object Detection", 8),
    ({"classifier free"}, "Generative Models", "Diffusion Models", 8),
    ({"parameter free"}, "Neural Networks", "Model Parameters", 8),
    ({"hessian free"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"model free"}, "Reinforcement Learning", "Learning Techniques", 8),

    # === X-wise (21 unique unclassified terms) ===
    ({"layer wise"}, "Neural Networks", "Core Models", 8),
    ({"channel wise"}, "Neural Networks", "Core Models", 8),
    ({"pixel wise"}, "Computer Vision", "Image Processing", 8),
    ({"token wise"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"element wise"}, "Neural Networks", "Core Models", 8),
    ({"point wise"}, "Computer Vision", "", 7),

    # === X-level (20 unique unclassified terms) ===
    ({"token level"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"sentence level"}, "Natural Language Processing", "Text Analysis", 8),
    ({"document level"}, "Natural Language Processing", "Text Analysis", 8),
    ({"word level"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"character level"}, "Natural Language Processing", "Text Preprocessing", 8),
    ({"system level"}, "AI Applications", "", 7),
    ({"dataset level"}, "Data Preprocessing", "Data Enhancement", 7),
    ({"concept level"}, "AI Theory", "Knowledge Representation", 7),

    # === X-specific (9 unique unclassified terms) ===
    ({"domain specific"}, "Transfer Learning", "Domain Adaptation", 8),
    ({"task specific"}, "Transfer Learning", "Model Adaptation", 8),
    ({"dataset specific"}, "Data Preprocessing", "Data Enhancement", 7),

    # === X-centric (6 unique unclassified terms) ===
    ({"data centric"}, "Data Preprocessing", "Data Enhancement", 8),
    ({"object centric"}, "Computer Vision", "Image Processing", 8),
    ({"user centric"}, "AI Applications", "", 7),
    ({"model centric"}, "Neural Networks", "Core Models", 7),

    # === X-only (3 unique unclassified terms) ===
    ({"decoder only"}, "Neural Networks", "Core Models", 8),
    ({"encoder only"}, "Neural Networks", "Core Models", 8),

    # === X-scale (4 unique unclassified terms) ===
    ({"large scale"}, "Neural Networks", "Training Techniques", 8),
    ({"multi scale"}, "Computer Vision", "Image Processing", 8),
    ({"brain scale"}, "Neural Networks", "Specialized Architectures", 7),

    # ==========================================
    # Batch 4 — Named acronyms and specific techniques
    # Extracted from the hardest 200 unclassified terms.
    # Single-token rules for distinct acronyms (SMOTE, UMAP, etc.)
    # and multi-word rules for specific named techniques.
    # ==========================================

    # === Named acronyms (high-confidence single-token patterns) ===
    ({"smote"}, "Data Preprocessing", "Data Enhancement", 9),
    ({"umap"}, "Dimensionality Reduction", "Feature Reduction", 9),
    ({"lime"}, "Explainable AI", "Model Interpretability", 9),
    ({"arima"}, "Statistical Methods", "Bayesian Inference", 9),
    ({"roc"}, "Model Evaluation", "Performance Metrics", 9),
    ({"hnsw"}, "Similarity & Deduplication", "", 8),
    ({"neat"}, "AI Applications", "", 8),
    ({"ffjord"}, "Generative Models", "", 8),
    ({"kfac", "k fac"}, "Optimization Algorithms", "Gradient-based Optimizers", 8),
    ({"adni"}, "AI Applications", "", 7),

    # === Named techniques (multi-word patterns) ===
    ({"contrastive divergence"}, "Neural Networks", "Training Techniques", 9),
    ({"variational information bottleneck"}, "Representation Learning", "Feature Learning", 9),
    ({"variational information"}, "Representation Learning", "Feature Learning", 8),
    ({"uniform manifold approximation"}, "Dimensionality Reduction", "Feature Reduction", 9),
    ({"random fourier features"}, "Neural Networks", "Core Models", 8),
    ({"random fourier"}, "Neural Networks", "Core Models", 8),
    ({"co occurrence matrix"}, "Representation Learning", "Vector Representations", 8),
    ({"progressive growing"}, "Generative Models", "Adversarial Networks", 8),
    ({"lipschitz continuity"}, "Neural Networks", "Regularization Techniques", 8),
    ({"lipschitz"}, "Neural Networks", "Regularization Techniques", 8),
    ({"orthogonal initialization"}, "Neural Networks", "Training Techniques", 8),
    ({"mean cancellation"}, "Neural Networks", "Training Techniques", 8),
    ({"cyclical consistency"}, "Computer Vision", "Image Processing", 8),
    ({"random projection"}, "Dimensionality Reduction", "Feature Reduction", 8),
    ({"covariance"}, "Statistical Methods", "Bayesian Inference", 8),
    ({"heteroscedastic"}, "Statistical Methods", "Bayesian Inference", 9),
    ({"boltzmann"}, "Neural Networks", "Core Models", 9),
]

# Batch 5 — Data processing and high-frequency multi-word techniques
BATCH_5 = [
    # === Single-token data/processing patterns (p8) ===
    ({"loss"}, "Loss Functions", "Classification Loss", 8),
    ({"transformation"}, "Data Preprocessing", "Data Scaling", 8),
    ({"mining"}, "Data Processing", "Data Pipeline", 8),
    ({"swap"}, "Data Preprocessing", "Data Enhancement", 8),
    # === Named multi-word techniques (p8) ===
    ({"mixture of experts"}, "Neural Networks", "Specialized Architectures", 8),
    ({"deep equilibrium"}, "Neural Networks", "Core Models", 8),
    ({"mean teacher"}, "Learning Paradigms", "Low-Data Learning", 8),
    ({"state space"}, "Statistical Methods", "Probabilistic Models", 8),
    ({"noise injection"}, "Neural Networks", "Regularization Techniques", 8),
    ({"association rule"}, "Data Processing", "Data Pipeline", 8),
    ({"frozen pretrained"}, "Transfer Learning", "Model Adaptation", 8),
    ({"mutual information"}, "Representation Learning", "Feature Learning", 8),
    ({"autoregressive"}, "Statistical Methods", "Probabilistic Models", 8),
    ({"dynamic loss"}, "Loss Functions", "Classification Loss", 8),
    ({"polyak averaging"}, "Optimization Algorithms", "Gradient-Based Optimizers", 8),
    ({"log ratio"}, "Data Preprocessing", "Data Scaling", 8),
    ({"information maximization"}, "Representation Learning", "Feature Learning", 8),
    ({"exponential family"}, "Statistical Methods", "Probabilistic Models", 8),
    ({"cycle consistency"}, "Computer Vision", "Image Processing", 8),
]
AUTO_CLASSIFICATION_RULES.extend(BATCH_5)

# Batch 6 — Named acronyms from the hard-200 analysis
BATCH_6 = [
    # === Named acronyms (3-4+ chars, high domain specificity) ===
    ({"nms"}, "Computer Vision", "Object Detection", 9),
    ({"cart"}, "Learning Paradigms", "Supervised Learning", 9),
    ({"slam"}, "Computer Vision", "Image Processing", 8),
    ({"admm"}, "Optimization Algorithms", "Gradient-Based Optimizers", 8),
    ({"focl"}, "Machine Learning Frameworks", "Automated Processes", 8),
    ({"ecoc"}, "Loss Functions", "Classification Loss", 8),
    ({"slim"}, "Model Optimization", "Model Compression", 8),
    ({"nmf"}, "Dimensionality Reduction", "Feature Reduction", 8),
    ({"nist"}, "Ethics & Governance", "", 8),
    ({"wer"}, "Model Evaluation", "Performance Metrics", 8),
    ({"cer"}, "Model Evaluation", "Performance Metrics", 8),
    ({"psnr"}, "Computer Vision", "Image Processing", 8),
    ({"mpnn"}, "Neural Networks", "Graph-Based Models", 8),
    ({"pos"}, "Natural Language Processing", "Text Analysis", 8),
    # === Named multi-word concepts (domain-specific phrases) ===
    ({"theory of mind"}, "AI Theory", "Knowledge Representation", 9),
    ({"human in the loop"}, "AI Applications", "", 9),
    ({"society of mind"}, "AI Theory", "Knowledge Representation", 8),
    ({"kernel trick"}, "Statistical Methods", "Probabilistic Models", 8),
    ({"bloom filter"}, "Data Processing", "Data Pipeline", 8),
    ({"ab testing", "a/b testing"}, "Model Evaluation", "Performance Metrics", 8),
    ({"approximate nearest neighbor"}, "Similarity & Deduplication", "", 8),
    ({"composition of experts"}, "Neural Networks", "Specialized Architectures", 8),
    # === Moderate-confidence single-token patterns ===
    ({"calibration"}, "Model Evaluation", "Performance Metrics", 8),
    ({"sequence"}, "Natural Language Processing", "Language Modeling", 8),
    ({"decomposition"}, "Dimensionality Reduction", "Feature Reduction", 8),
    ({"decoding"}, "Natural Language Processing", "Language Generation", 8),
    # === Named architecture patterns ===
    ({"bottleneck"}, "Representation Learning", "Feature Learning", 8),
    ({"projection"}, "Dimensionality Reduction", "Feature Reduction", 8),
    ({"backbone"}, "Neural Networks", "Core Models", 8),
]
AUTO_CLASSIFICATION_RULES.extend(BATCH_6)

# ── Study-family taxonomy mapping (metadata-driven fallback) ──────

STUDY_FAMILY_TAXONOMY_MAP: dict[str, tuple[str, str]] = {
    "Evaluation": ("Model Evaluation", "Performance Metrics"),
    "Ethics & Governance": ("Ethics & Governance", "Fairness in ML"),
    "Statistics": ("Statistical Methods", "Bayesian Inference"),
    "Similarity & Deduplication": ("Similarity & Deduplication", ""),
}


def family_content_profile(family: str) -> dict[str, str]:
    return FAMILY_CONTENT_PROFILES.get(
        family,
        {
            "note": "Watch how it fits relative to nearby concepts in the learning graph.",
            "boundary": "Ask how it differs from nearby concepts and what changes when the term is applied.",
            "quiz_anchor": "Use the broader learning graph to situate it before you memorize the wording.",
            "compare_anchor": "Compare it with nearby concepts so the boundary stays visible.",
        },
    )


def _sanitize_definition_title(title: str) -> str:
    title = title.strip()
    while title.startswith(("*", "-", "•")):
        title = title[1:].strip()
    while title.endswith("*"):
        title = title[:-1].strip()
    return title


def _is_rejectable_definition_title(title: str, title_key: str) -> bool:
    if not title_key or not title:
        return True
    lower_title = title_key.lower()
    reject_prefixes = {"new additions", "terms", "term", "v", "section", "topics", "topic"}
    if lower_title in reject_prefixes:
        return True
    if lower_title.startswith("here are") or lower_title.startswith("here's") or lower_title.startswith("this is"):
        return True
    return False


def _is_valid_definition_body(body: str, title: str) -> bool:
    if not body or len(body) < DEFINITION_BODY_MIN_CHARS:
        return False
    if len(body) > 4000:
        return False
    if title.lower() == body.lower():
        return False
    if not re.search(r"[a-zA-Z]", body):
        return False
    return True


def _normalize_title_for_match(value: str) -> str:
    if len(value) > DEFINITION_TITLE_MAX_CHARS:
        return value[:DEFINITION_TITLE_MAX_CHARS].strip()
    return value


def _parse_definition_row(raw_value: str) -> tuple[str, str] | None:
    if not raw_value or len(raw_value) > 5000:
        return None

    match = DEFINITION_NUMBERED_RE.match(raw_value)
    if match:
        title = normalize_label(match.group("numbered_title") or match.group("numbered_title_plain") or "")
        title = _normalize_title_for_match(_sanitize_definition_title(title))
        body = normalize_label(match.group("body"))
        return title, body

    if ":" in raw_value:
        match = DEFINITION_FALLBACK_RE.match(raw_value)
        if not match:
            return None

        title = _normalize_title_for_match(_sanitize_definition_title(match.group("title")))
        body = normalize_label(match.group("body"))
        return title, body

    return None


PAREN_ALIAS_RE = re.compile(r"^(?P<base>.+?)\s+\((?P<paren>[^()]+)\)\s*$")


def normalize_label(value: object | None) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u2011", "-").replace("\u2013", "-").replace("\u2014", "-")
    text = WHITESPACE_RE.sub(" ", text.strip())
    return text


def normalize_key(value: object | None) -> str:
    return normalize_label(value).lower()


def slugify(value: object | None) -> str:
    slug = NON_ALNUM_RE.sub("-", normalize_key(value)).strip("-")
    return slug or "untitled"


def semantic_tokens(value: object | None) -> list[str]:
    text = normalize_label(value).lower()
    replacements = {
        "&": " and ",
        "+": " plus ",
        "#": " sharp ",
        "/": " ",
        "\\": " ",
        "_": " ",
        "-": " ",
        "(": " ",
        ")": " ",
        "[": " ",
        "]": " ",
        "{": " ",
        "}": " ",
        ",": " ",
        ":": " ",
        ";": " ",
        ".": " ",
        "'": "",
        '"': " ",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return TOKEN_RE.findall(text)


def semantic_key(value: object | None) -> str:
    tokens = semantic_tokens(value)
    return " ".join(tokens) if tokens else normalize_key(value)


def title_tokens_for_graph(value: object | None) -> list[str]:
    stopwords = {
        "and", "or", "of", "for", "the", "in", "to", "with", "a", "an", "vs", "on",
        "architecture", "architectures", "model", "models", "network", "networks",
        "optimizer", "optimizers", "algorithm", "algorithms", "method", "methods",
        "function", "functions", "system", "systems",
    }
    deduped: list[str] = []
    seen: set[str] = set()
    for token in semantic_tokens(value):
        if token in stopwords or len(token) <= 1 or token in seen:
            continue
        seen.add(token)
        deduped.append(token)
    return deduped


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=True, separators=(",", ":")) + "\n", encoding="utf-8")


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def collect_artifact_hashes(out_dir: Path) -> dict[str, str]:
    hashes: dict[str, str] = {}
    for path in sorted(out_dir.rglob("*.json")):
        relative = path.relative_to(out_dir).as_posix()
        if relative == "manifest.json":
            continue
        hashes[relative] = file_sha256(path)
    return hashes


def replace_directory_after_success(staging_dir: Path, final_dir: Path) -> None:
    final_dir.parent.mkdir(parents=True, exist_ok=True)
    previous_dir = final_dir.with_name(f".{final_dir.name}.previous")
    if previous_dir.exists():
        shutil.rmtree(previous_dir)
    if final_dir.exists():
        os.replace(final_dir, previous_dir)
    try:
        os.replace(staging_dir, final_dir)
    except Exception:
        if previous_dir.exists() and not final_dir.exists():
            os.replace(previous_dir, final_dir)
        raise
    if previous_dir.exists():
        shutil.rmtree(previous_dir)


def resolve_slug(base_slug: str, used_slugs: Counter[str]) -> str:
    if used_slugs[base_slug] == 0:
        used_slugs[base_slug] += 1
        return base_slug

    used_slugs[base_slug] += 1
    return f"{base_slug}-{used_slugs[base_slug]}"


def resolve_shard_id(slug: str) -> str:
    alnum = "".join(character for character in slug.lower() if character.isalnum())
    if not alnum:
        return "misc"
    if len(alnum) == 1:
        return f"{alnum}_"
    return alnum[:2]


def _is_rejectable_inventory_title(title: str) -> bool:
    key = normalize_key(title)
    if not key or key in INVENTORY_REJECT_KEYS:
        return True
    return not re.search(r"[a-zA-Z]", title)


def extract_term_inventory(glossary_path: Path) -> list[dict]:
    workbook = load_workbook(glossary_path, read_only=True, data_only=True)
    sheet = workbook["main"]

    terms_by_key: dict[str, dict] = {}

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        for column_index in INVENTORY_COLUMNS:
            value = row[column_index - 1] if len(row) >= column_index else None
            title = normalize_label(value)
            if not title or _is_rejectable_inventory_title(title):
                continue

            title_key = normalize_key(title)
            entry = terms_by_key.get(title_key)
            source_cell = {
                "rowNumber": row_number,
                "columnIndex": column_index,
                "columnLetter": chr(64 + column_index),
            }
            if entry is None:
                terms_by_key[title_key] = {
                    "title": title,
                    "titleKey": title_key,
                    "semanticKey": semantic_key(title),
                    "rowNumber": row_number,
                    "inventoryRows": [row_number],
                    "inventoryCells": [f"{source_cell['columnLetter']}{row_number}"],
                    "inventoryColumns": [source_cell["columnLetter"]],
                    "sourceCells": [source_cell],
                }
                continue

            entry.setdefault("inventoryRows", []).append(row_number)
            entry.setdefault("inventoryCells", []).append(f"{source_cell['columnLetter']}{row_number}")
            entry.setdefault("inventoryColumns", []).append(source_cell["columnLetter"])
            entry.setdefault("sourceCells", []).append(source_cell)
            if row_number < entry["rowNumber"]:
                entry["rowNumber"] = row_number

    terms = []
    for term in terms_by_key.values():
        term["inventoryRows"] = sorted(set(term.get("inventoryRows", [])))
        term["inventoryCells"] = sorted(set(term.get("inventoryCells", [])))
        term["inventoryColumns"] = sorted(set(term.get("inventoryColumns", [])))
        term["sourceCells"] = sorted(
            term.get("sourceCells", []), key=lambda item: (item["rowNumber"], item["columnIndex"])
        )
        terms.append(term)

    return sorted(terms, key=lambda item: item["rowNumber"])


def score_title_quality(term: dict) -> tuple[int, int, int, int]:
    title = term["title"]
    words = [piece for piece in re.split(r"\s+", title) if piece]
    uppercase_word_count = sum(1 for word in words if word[:1].isupper())
    punctuation_signal = sum(title.count(character) for character in "-()/")
    return (
        uppercase_word_count,
        punctuation_signal,
        len(title),
        -term["rowNumber"],
    )


def merge_semantic_variants(terms: list[dict]) -> tuple[list[dict], list[dict]]:
    groups: dict[str, list[dict]] = defaultdict(list)
    for term in terms:
        groups[term["semanticKey"]].append(term)

    merged_terms: list[dict] = []
    canonicalization_groups: list[dict] = []

    for semantic_group_key, group in groups.items():
        if len(group) == 1:
            term = dict(group[0])
            term["titleKeys"] = [term["titleKey"]]
            term["variantTitles"] = []
            term["inventoryRows"] = sorted(set(term.get("inventoryRows", [term["rowNumber"]])))
            term["inventoryCells"] = sorted(set(term.get("inventoryCells", [])))
            term["inventoryColumns"] = sorted(set(term.get("inventoryColumns", [])))
            term["sourceCells"] = sorted(
                term.get("sourceCells", []), key=lambda item: (item["rowNumber"], item["columnIndex"])
            )
            merged_terms.append(term)
            continue

        canonical = dict(max(group, key=score_title_quality))
        aliases = []
        title_keys = []
        source_rows = []
        source_cells = []
        source_cells_seen: set[tuple[int, int]] = set()
        inventory_cells: set[str] = set()
        inventory_columns: set[str] = set()

        for variant in sorted(group, key=lambda item: item["rowNumber"]):
            title_keys.append(variant["titleKey"])
            source_rows.extend(variant.get("inventoryRows", [variant["rowNumber"]]))
            source_rows.extend(variant.get("duplicateRows", []))
            inventory_cells.update(variant.get("inventoryCells", []))
            inventory_columns.update(variant.get("inventoryColumns", []))
            for cell in variant.get("sourceCells", []):
                cell_key = (cell["rowNumber"], cell["columnIndex"])
                if cell_key in source_cells_seen:
                    continue
                source_cells_seen.add(cell_key)
                source_cells.append(cell)
            if normalize_key(variant["title"]) != normalize_key(canonical["title"]):
                aliases.append(variant["title"])

        canonical["titleKeys"] = sorted(set(title_keys))
        canonical["variantTitles"] = sorted(set(aliases), key=str.lower)
        canonical["duplicateRows"] = sorted(set(source_rows) - {canonical["rowNumber"]})
        canonical["inventoryRows"] = sorted(set(source_rows))
        canonical["inventoryCells"] = sorted(inventory_cells)
        canonical["inventoryColumns"] = sorted(inventory_columns)
        canonical["sourceCells"] = sorted(source_cells, key=lambda item: (item["rowNumber"], item["columnIndex"]))
        merged_terms.append(canonical)

        canonicalization_groups.append(
            {
                "semanticKey": semantic_group_key,
                "canonicalTitle": canonical["title"],
                "canonicalRow": canonical["rowNumber"],
                "variantTitles": [
                    {
                        "title": variant["title"],
                        "rowNumber": variant["rowNumber"],
                    }
                    for variant in sorted(group, key=lambda item: item["rowNumber"])
                    if variant["title"] != canonical["title"]
                ],
                "mergedRowCount": len(sorted(set(source_rows))),
            }
        )

    return sorted(merged_terms, key=lambda item: item["rowNumber"]), canonicalization_groups


def extract_taxonomy_lookup(glossary_path: Path) -> dict[str, dict]:
    workbook = load_workbook(glossary_path, read_only=True, data_only=True)
    sheet = workbook["main"]

    taxonomy: dict[str, dict] = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        term = normalize_label(row[13] if len(row) > 13 else None)
        category = normalize_label(row[14] if len(row) > 14 else None)
        sub_category = normalize_label(row[15] if len(row) > 15 else None)
        if not term or normalize_key(term) == "topic":
            continue

        term_key = normalize_key(term)
        existing = taxonomy.get(term_key)
        candidate = {
            "term": term,
            "category": category,
            "subCategory": sub_category,
            "rowNumber": row_number,
        }
        if existing is None or (not existing["category"] and category):
            taxonomy[term_key] = candidate

    return taxonomy


def load_taxonomy_registry(registry_path: Path | None) -> dict[str, dict]:
    """Load the authoritative editorial taxonomy registry.

    The registry (data/taxonomy-registry.json, generated by
    tools/extract_taxonomy_registry.py) holds curated (category, subCategory)
    decisions keyed by normalized term title/alias. It is the primary taxonomy
    source so the build is reproducible from the editorial record rather than
    from a brittle derived keyword-rule layer. Returns {} when no path is given
    or the file is absent (build falls back to workbook + auto-rules).
    """
    if registry_path is None:
        return {}
    path = Path(registry_path)
    if not path.is_file():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    entries = payload.get("entries", {}) if isinstance(payload, dict) else {}
    registry: dict[str, dict] = {}
    for key, decision in entries.items():
        if not isinstance(decision, dict):
            continue
        category = decision.get("category")
        if not category:
            continue
        registry[key] = {
            "term": key,
            "category": category,
            "subCategory": decision.get("subCategory") or "",
            "rowNumber": -1,
        }
    return registry


def extract_definition_lookup(glossary_path: Path) -> dict[str, dict]:
    workbook = load_workbook(glossary_path, read_only=True, data_only=True)
    sheet = workbook["main"]

    definitions: dict[str, dict] = {}
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        raw_value = normalize_label(row[7] if len(row) > 7 else None)
        if not raw_value:
            continue

        parsed = _parse_definition_row(raw_value)
        if not parsed:
            continue

        title, body = parsed
        title = _normalize_title_for_match(title)
        title_key = normalize_key(title)
        if _is_rejectable_definition_title(title, title_key):
            continue
        if not _is_valid_definition_body(body, title):
            continue
        if title_key in definitions:
            if len(body) <= len(definitions[title_key]["body"]):
                continue

        definitions[title_key] = {
            "title": title,
            "body": body,
            "rowNumber": row_number,
        }

    return definitions


def extract_structure_registry(structure_path: Path) -> dict:
    workbook = load_workbook(structure_path, read_only=True, data_only=True)
    sheet = workbook["Sheet2"]

    header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    outline_rows = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        outline_label = normalize_label(row[0] if len(row) > 0 else None)
        if outline_label:
            outline_rows.append({"rowNumber": row_number, "label": outline_label})

    fields = []
    for index, value in enumerate(header_row, start=1):
        label = normalize_label(value)
        if not label:
            continue
        section = label.split(" – ", 1)[0].split(" - ", 1)[0].split(" — ", 1)[0]
        if section in STRUCTURE_LAUNCH_SECTIONS:
            layer = "launch-runtime"
        elif section in STRUCTURE_EDITORIAL_SECTIONS:
            layer = "editorial-expansion"
        else:
            layer = "backlog"
        fields.append(
            {
                "columnIndex": index,
                "columnKey": slugify(label),
                "label": label,
                "section": section,
                "layer": layer,
            }
        )

    section_groups: dict[str, dict] = {}
    for field in fields:
        section = field["section"]
        group = section_groups.setdefault(
            section,
            {
                "section": section,
                "layer": field["layer"],
                "fieldCount": 0,
                "firstColumnIndex": field["columnIndex"],
                "lastColumnIndex": field["columnIndex"],
                "sampleFields": [],
            },
        )
        group["fieldCount"] += 1
        group["firstColumnIndex"] = min(group["firstColumnIndex"], field["columnIndex"])
        group["lastColumnIndex"] = max(group["lastColumnIndex"], field["columnIndex"])
        if len(group["sampleFields"]) < 3:
            group["sampleFields"].append(field["label"])

    return {
        "sheetName": "Sheet2",
        "fieldCount": len(fields),
        "fields": fields,
        "outline": outline_rows,
        "layerCounts": {
            "launch-runtime": sum(1 for field in fields if field["layer"] == "launch-runtime"),
            "editorial-expansion": sum(1 for field in fields if field["layer"] == "editorial-expansion"),
            "backlog": sum(1 for field in fields if field["layer"] == "backlog"),
        },
        "sectionGroups": sorted(
            section_groups.values(),
            key=lambda item: (item["layer"], item["firstColumnIndex"], item["section"].lower()),
        ),
        "launchSections": list(STRUCTURE_LAUNCH_SECTION_ORDER),
        "editorialSections": list(STRUCTURE_EDITORIAL_SECTION_ORDER),
    }

def compact_title_list(values: list[str], limit: int = 2) -> str:
    if not values:
        return ""
    selected = values[:limit]
    if len(selected) == 1:
        return selected[0]
    if len(selected) == 2:
        return f"{selected[0]} and {selected[1]}"
    return f"{', '.join(selected[:-1])}, and {selected[-1]}"


def family_study_note(term: dict, taxonomy: dict | None) -> str:
    family = infer_study_family(term, taxonomy)
    profile = family_content_profile(family)
    if family in FAMILY_CONTENT_PROFILES:
        return profile["note"]

    if taxonomy and taxonomy["category"] and taxonomy["subCategory"]:
        return (
            f"Watch how it fits inside {taxonomy['category']} / {taxonomy['subCategory']} and how the neighboring terms change its meaning."
        )
    if taxonomy and taxonomy["category"]:
        return f"Watch how it fits inside {taxonomy['category']} and how nearby concepts narrow or extend the idea."
    return "Watch how it behaves relative to nearby concepts in the learning graph."


def build_comparison_items(term: dict, links: dict[str, list[str]]) -> list[str]:
    items: list[str] = []
    if links["prerequisites"]:
        items.append(f"Before this term, get comfortable with {compact_title_list(links['prerequisites'])}.")
    if links["related"]:
        items.append(f"Compare it with {compact_title_list(links['related'])} to separate overlap from difference.")
    if links["alternatives"]:
        items.append(f"Treat {compact_title_list(links['alternatives'])} as neighboring labels until you confirm the source usage.")
    if links["next"]:
        items.append(f"After this term, continue into {compact_title_list(links['next'])}.")

    if not items:
        items.append("Use the taxonomy and the concept graph to place this term before you try to memorize a one-line definition.")

    return items[:4]


def build_recall_steps(term: dict, taxonomy: dict | None, links: dict[str, list[str]]) -> list[dict]:
    steps = [
        {
            "label": "Say it simply",
            "body": f"Rewrite {term['title']} in one sentence before trying to preserve the source wording.",
        },
        {
            "label": "Place it",
            "body": family_study_note(term, taxonomy),
        },
        {
            "label": "Link it",
            "body": (
                f"Connect it to {compact_title_list(links['related']) or 'a nearby neighbor'} and one downstream idea from the graph."
            ),
        },
    ]

    if links["prerequisites"]:
        steps.append(
            {
                "label": "Review first",
                "body": f"Check {compact_title_list(links['prerequisites'])} so you can see what this term assumes.",
            }
        )
    else:
        steps.append(
            {
                "label": "Review first",
                "body": "If the term feels abstract, look at one nearby concept before returning to the main definition.",
            }
        )

    return steps


def first_or_empty(values: list[str]) -> str:
    return values[0] if values else ""


def build_at_a_glance_rows(term: dict, summary: str, taxonomy: dict | None, links: dict[str, list[str]]) -> list[dict]:
    family = taxonomy["category"] if taxonomy and taxonomy["category"] else infer_study_family(term, taxonomy) or "Unclassified"
    subfamily = taxonomy["subCategory"] if taxonomy and taxonomy["subCategory"] else "Unclassified"
    profile = family_content_profile(family)
    rows = [
        {"label": "Family", "value": f"{family} / {subfamily}" if subfamily != "Unclassified" else family},
        {
            "label": "Best first read",
            "value": compact_title_list(links["prerequisites"]) if links["prerequisites"] else "Use the overview block",
        },
        {
            "label": "Best comparison",
            "value": compact_title_list(links["related"]) if links["related"] else "Use the connection map",
        },
        {
            "label": "Next step",
            "value": compact_title_list(links["next"]) if links["next"] else "Use the recall drill",
        },
        {
            "label": "Remember",
            "value": profile["note"],
        },
    ]

    return rows


def build_quiz_block(term: dict, taxonomy: dict | None, links: dict[str, list[str]]) -> dict:
    question: str
    options: list[str]
    answer_index: int
    explanation: str

    prerequisite = first_or_empty(links["prerequisites"])
    related = first_or_empty(links["related"])
    next_term = first_or_empty(links["next"])

    if prerequisite:
        question = f"What should you review first before studying {term['title']}?"
        options = [
            f"Review {prerequisite}",
            f"Jump to {related or next_term or 'a nearby concept'}",
            "Memorize the title only",
            "Skip straight to the next topic",
        ]
        answer_index = 0
        explanation = (
            f"{term['title']} sits after {prerequisite}, so that prerequisite gives you the assumptions and vocabulary"
            " you need before the deeper block stack makes sense."
        )
    elif related:
        question = f"How should you place {term['title']} in the learning graph?"
        options = [
            "Treat it as an isolated fact",
            f"Compare it with {related}",
            "Ignore the taxonomy",
            f"Jump only to {next_term or 'the next term'}",
        ]
        answer_index = 1
        explanation = (
            f"{term['title']} is easiest to understand by comparing it with {related}, since neighbors expose"
            " overlap, difference, and boundary conditions."
        )
    else:
        family = taxonomy["category"] if taxonomy and taxonomy["category"] else infer_study_family(term, taxonomy) or "the nearby graph"
        profile = family_content_profile(family)
        question = f"What is the safest way to place {term['title']}?"
        options = [
            profile["quiz_anchor"],
            "Memorize the title only",
            "Treat it as unrelated to the graph",
            "Skip the overview and jump to exports",
        ]
        answer_index = 0
        explanation = (
            f"When specific links are sparse, the best anchor is the broader {family} context plus the overview block, comparison view, and recall drill."
        )

    return {
        "id": "quick-quiz",
        "type": "quiz",
        "title": "Quick Check",
        "question": question,
        "options": options,
        "answerIndex": answer_index,
        "explanation": explanation,
    }


def build_diagram_block(term: dict, taxonomy: dict | None, links: dict[str, list[str]]) -> dict:
    family = ""
    if taxonomy and taxonomy["category"]:
        family = taxonomy["category"]
        if taxonomy["subCategory"]:
            family = f"{family} / {taxonomy['subCategory']}"
    else:
        family = infer_study_family(term, taxonomy) or "the nearby graph"

    compare_items = dedupe_titles(
        [
            *links["related"],
            *links["alternatives"],
        ],
        3,
    )

    lanes = [
        {
            "label": "Before",
            "tone": "accent",
            "items": links["prerequisites"][:3],
        },
        {
            "label": "Compare",
            "tone": "secondary",
            "items": compare_items,
        },
        {
            "label": "After",
            "tone": "gold",
            "items": links["next"][:3],
        },
    ]

    return {
        "id": "concept-map",
        "type": "diagram",
        "title": "Concept Map",
        "center": {
            "label": term["title"],
            "caption": family,
        },
        "lanes": lanes,
        "takeaway": (
            f"Read {term['title']} as a node inside its learning graph: predecessor concepts give context, "
            "neighbors sharpen comparison, and next concepts show where the topic expands."
        ),
    }


def build_faq_block(term: dict, summary: str, taxonomy: dict | None, links: dict[str, list[str]]) -> dict:
    prerequisites = links["prerequisites"][:2]
    related = links["related"][:2]
    next_terms = links["next"][:2]
    family = taxonomy["category"] if taxonomy and taxonomy["category"] else infer_study_family(term, taxonomy) or "the surrounding graph"
    profile = family_content_profile(family)

    questions_and_answers = [
        {
            "question": f"What is {term['title']} in one sentence?",
            "answer": summary,
        },
        {
            "question": f"What should I know before studying {term['title']}?",
            "answer": (
                f"Start with {compact_title_list(prerequisites)}."
                if prerequisites
                else f"Use the broader {family} context, the overview block, and the comparison view to orient yourself. {profile['boundary']}"
            ),
        },
        {
            "question": f"What should I compare {term['title']} with?",
            "answer": (
                f"Compare it with {compact_title_list(related)} before moving on to {compact_title_list(next_terms)}."
                if related or next_terms
                else f"Compare it with the nearest concepts in {family}. {profile['compare_anchor']}"
            ),
        },
    ]

    return {
        "id": "quick-faq",
        "type": "faq",
        "title": "Quick FAQ",
        "items": questions_and_answers,
    }


def term_content_score(term: dict, taxonomy: dict | None, definition: dict | None, links: dict[str, list[str]]) -> int:
    link_count = sum(len(links[key]) for key in ("prerequisites", "related", "alternatives", "next"))
    score = 0
    if definition:
        score += 5
    if taxonomy and taxonomy["category"]:
        score += 2
    if taxonomy and taxonomy["subCategory"]:
        score += 1
    score += min(4, link_count)
    score += min(2, len(term.get("variantTitles", [])))
    if term.get("duplicateRows"):
        score += 1
    return score


def term_content_tier(term: dict, taxonomy: dict | None, definition: dict | None, links: dict[str, list[str]]) -> str:
    score = term_content_score(term, taxonomy, definition, links)
    if score >= 9:
        return "featured"
    if score >= 4:
        return "standard"
    return "sparse"


def build_deep_dive_block(term: dict, summary: str, taxonomy: dict | None, links: dict[str, list[str]]) -> dict | None:
    if not summary:
        return None

    family = taxonomy["category"] if taxonomy and taxonomy["category"] else infer_study_family(term, taxonomy) or "the nearby graph"
    profile = family_content_profile(family)
    subfamily = taxonomy["subCategory"] if taxonomy and taxonomy["subCategory"] else ""
    related = compact_title_list(links["related"]) if links["related"] else ""
    alternatives = compact_title_list(links["alternatives"]) if links["alternatives"] else ""
    prerequisites = compact_title_list(links["prerequisites"]) if links["prerequisites"] else ""
    next_terms = compact_title_list(links["next"]) if links["next"] else ""

    panels = [
        {
            "label": "Core idea",
            "tone": "accent",
            "body": summary.split(". ")[0].rstrip(".") + ".",
        },
        {
            "label": "Place it",
            "tone": "secondary",
            "body": f"Read it inside {family}{f' / {subfamily}' if subfamily else ''} before you try to memorize the wording.",
        },
        {
            "label": "Compare it",
            "tone": "gold",
            "body": (
                f"{profile['compare_anchor']} Use {related or alternatives or 'the nearby graph'} to see what this term is and what it is not."
            ),
        },
        {
            "label": "Move next",
            "tone": "accent",
            "body": (
                f"Start with {prerequisites or 'the overview block'}, then move into {next_terms or 'the next connected concept'}."
            ),
        },
        {
            "label": "Memory hook",
            "tone": "secondary",
            "body": family_study_note(term, taxonomy),
        },
    ]

    return {
        "id": "deep-dive",
        "type": "deep-dive",
        "title": "Featured Deep Dive",
        "panels": panels,
        "takeaway": (
            f"{term['title']} deserves a deeper treatment because the surrounding graph gives enough signal to teach it as a concept, not just as a definition."
        ),
    }


def build_comparison_block(term: dict, summary: str, taxonomy: dict | None, links: dict[str, list[str]]) -> dict:
    family = taxonomy["category"] if taxonomy and taxonomy["category"] else infer_study_family(term, taxonomy) or "the nearby graph"
    profile = family_content_profile(family)
    related = compact_title_list(links["related"]) if links["related"] else ""
    alternatives = compact_title_list(links["alternatives"]) if links["alternatives"] else ""
    prerequisites = compact_title_list(links["prerequisites"]) if links["prerequisites"] else ""
    next_terms = compact_title_list(links["next"]) if links["next"] else ""

    panels = [
        {
            "label": "What it is",
            "tone": "accent",
            "body": summary.split(". ")[0].rstrip(".") + ".",
        },
        {
            "label": "What it is not",
            "tone": "secondary",
            "body": f"It is not just {alternatives or related or 'an isolated vocabulary item'}; the graph shows how it sits inside {family}.",
        },
        {
            "label": "Common confusion",
            "tone": "gold",
            "body": (
                f"{profile['boundary']} Compare it against {related or alternatives or 'nearby concepts'} so the boundary stays visible instead of collapsing into a synonym pile."
            ),
        },
        {
            "label": "Best next step",
            "tone": "accent",
            "body": (
                f"Review {prerequisites or 'the overview block'} first, then move toward {next_terms or 'the next connected concept'}."
            ),
        },
    ]

    return {
        "id": "comparison",
        "type": "comparison",
        "title": "Compare and Contrast",
        "panels": panels,
        "takeaway": f"The comparison view keeps {term['title']} honest by showing the nearest neighbors, the real boundary, and the next study move.",
    }


def infer_study_family(term: dict, taxonomy: dict | None) -> str:
    if taxonomy and taxonomy["category"]:
        return taxonomy["category"]

    title_tokens = set(title_tokens_for_graph(term["title"]))
    family_rules = [
        ("Similarity & Deduplication", {"duplicate", "dedup", "deduplication", "similarity", "minhash", "shingle", "hash"}),
        ("Computer Vision", {"vision", "image", "video", "segmentation", "scene", "pose", "gaussian", "splatting", "3d", "ocr", "object"}),
        ("Natural Language Processing", {"language", "text", "token", "embedding", "translation", "generation", "rag", "nlp", "llm", "conversation", "retrieval"}),
        ("Reinforcement Learning", {"reinforcement", "policy", "agent", "bandit", "game", "exploration", "control"}),
        ("Statistics", {"bayesian", "probability", "statistical", "statistics", "inference", "causal"}),
        ("Evaluation", {"evaluation", "metric", "metrics", "benchmark", "accuracy", "precision", "recall", "f1"}),
        ("Ethics & Governance", {"ethics", "safety", "governance", "fairness", "privacy", "security"}),
        ("Neural Networks", {"optimizer", "loss", "gradient", "activation", "backprop", "neural", "network", "attention", "transformer", "layer", "encoder", "decoder"}),
    ]
    for family, keywords in family_rules:
        if title_tokens & keywords:
            return family

    return ""


def auto_classify_term(
    term_titles: list[str],
    taxonomy_lookup: dict[str, dict],
    auto_rules: list[tuple[set[str], str, str, int]],
    threshold: int = 8,
) -> tuple[str, str, bool]:
    """Try to classify a term using AUTO_CLASSIFICATION_RULES.

    Returns (category, subcategory, was_auto_classified).
    If no rule matches at or above the threshold, returns ("", "", False).
    """
    
    # Noise particles that appear as variant-suffix noise in unclassified terms.
    # Strip using word boundaries to avoid false positives like 'variants' in 'invariants'.
    _NOISE_RE = re.compile(r"\b(?:extensions|techniques|enhancements|variants)\b")
    best_score = 0
    best_category = ""
    best_subcategory = ""

    for title in term_titles:
        lower_title = title.lower().strip().replace("-", " ")
        lower_title = _NOISE_RE.sub(" ", lower_title)
        lower_title = re.sub(r"\s+", " ", lower_title).strip()
        tokens = set(semantic_tokens(title))

        for keywords, category, subcategory, priority in auto_rules:
            multi_word_matches = 0
            single_word_matches = 0

            for keyword in keywords:
                if " " in keyword:
                    if keyword in lower_title:
                        multi_word_matches += 1
                else:
                    if keyword in tokens:
                        single_word_matches += 1

            total_matches = multi_word_matches + single_word_matches
            if total_matches == 0:
                continue

            score = (multi_word_matches * 3 + single_word_matches) * priority
            if score > best_score:
                best_score = score
                best_category = category
                best_subcategory = subcategory

    if best_score >= threshold:
        return best_category, best_subcategory, True
    return "", "", False


def build_summary(term: dict, definition: dict | None, taxonomy: dict | None, links: dict[str, list[str]]) -> str:
    summary = definition["body"] if definition else ""
    if not summary and taxonomy and taxonomy["category"] and taxonomy["subCategory"]:
        summary = (
            f"{term['title']} belongs to the {taxonomy['category']} area of AI/ML, "
            f"within {taxonomy['subCategory']}."
        )
    elif not summary and taxonomy and taxonomy["category"]:
        summary = f"{term['title']} belongs to the {taxonomy['category']} area of AI/ML."
    elif not summary:
        summary = f"{term['title']} is a glossary node in the AI/ML learning graph."

    detail_parts: list[str] = []
    if taxonomy and taxonomy["category"]:
        if taxonomy["subCategory"]:
            detail_parts.append(f"It sits inside {taxonomy['category']} / {taxonomy['subCategory']}.")
        else:
            detail_parts.append(f"It sits inside {taxonomy['category']}.")

    if links["prerequisites"]:
        detail_parts.append(f"Start after {compact_title_list(links['prerequisites'])}.")
    elif links["related"]:
        detail_parts.append(f"Compare it with {compact_title_list(links['related'])}.")

    if links["next"]:
        detail_parts.append(f"It opens into {compact_title_list(links['next'])}.")

    if detail_parts:
        summary = f"{summary.rstrip()} {' '.join(detail_parts)}"

    return summary


def build_curriculum_map_block(launch_contract: dict) -> dict:
    return {
        "id": "curriculum-map",
        "type": "curriculum-map",
        "title": "Curriculum Map",
        "sections": launch_contract["launchSections"],
    }


def build_structure_expansion_block(structure_registry: dict) -> dict:
    sections = [
        section
        for section in structure_registry["sectionGroups"]
        if section["layer"] != "launch-runtime"
    ]
    sections.sort(
        key=lambda item: (
            STRUCTURE_LAYER_PRIORITY.get(item["layer"], len(STRUCTURE_LAYER_PRIORITY)),
            -item["fieldCount"],
            item["section"].lower(),
        )
    )
    highlighted_sections = sections[:8]
    return {
        "id": "structure-expansion",
        "type": "structure-expansion",
        "title": "Structure Expansion",
        "sections": highlighted_sections,
    }


def build_blocks(
    term: dict,
    summary: str,
    definition: dict | None,
    taxonomy: dict | None,
    links: dict[str, list[str]],
    launch_contract: dict,
    structure_registry: dict,
    editorial_tier: str,
) -> list[dict]:
    taxonomy_items = []
    if taxonomy:
        if taxonomy["category"]:
            taxonomy_items.append(f"Category: {taxonomy['category']}")
        if taxonomy["subCategory"]:
            taxonomy_items.append(f"Sub-category: {taxonomy['subCategory']}")
    if not taxonomy_items:
        taxonomy_items.append("Classification pending in the source taxonomy lookup.")

    connection_items = []
    if links["prerequisites"]:
        connection_items.append(f"Prerequisites: {', '.join(links['prerequisites'])}")
    if links["related"]:
        connection_items.append(f"Related: {', '.join(links['related'])}")
    if links["alternatives"]:
        connection_items.append(f"Alternatives: {', '.join(links['alternatives'])}")
    if links["next"]:
        connection_items.append(f"Next: {', '.join(links['next'])}")
    if not connection_items:
        connection_items.append("No promoted graph neighbors were available from the current import.")

    study_steps = []
    if taxonomy and taxonomy["category"]:
        if taxonomy["subCategory"]:
            study_steps.append(
                {
                    "label": "Situate it",
                    "body": f"Place this term inside {taxonomy['category']} / {taxonomy['subCategory']} before you memorize the wording.",
                }
            )
        else:
            study_steps.append(
                {
                    "label": "Situate it",
                    "body": f"Place this term inside {taxonomy['category']} before you memorize the wording.",
                }
            )
    else:
        study_steps.append(
            {
                "label": "Find nearby concepts",
                "body": "Locate the closest neighboring concepts before memorizing a phrasing.",
            }
        )

    study_steps.extend(
        [
            {
                "label": "Define it",
                "body": "Write the term in your own words before memorizing a textbook phrasing.",
            },
            {
                "label": "Trace the graph",
                "body": "Follow one prerequisite and one downstream concept from the learning graph.",
            },
            {
                "label": "Record an example",
                "body": "Capture an example, use case, or implementation note in your personal notebook.",
            },
        ]
    )

    blocks = [
        {
            "id": "overview",
            "type": "markdown",
            "title": "Overview",
            "body": summary,
        },
        {
            "id": "taxonomy",
            "type": "bullets",
            "title": "Field Position",
            "items": taxonomy_items,
        },
        {
            "id": "connections",
            "type": "bullets",
            "title": "Connection Map",
            "items": connection_items,
        },
        {
            "id": "study-prompts",
            "type": "steps",
            "title": "Study Prompts",
            "steps": study_steps,
        },
        {
            "id": "why-it-matters",
            "type": "markdown",
            "title": "Why It Matters",
            "body": family_study_note(term, taxonomy),
        },
        {
            "id": "comparison-notes",
            "type": "bullets",
            "title": "Comparison Notes",
            "items": build_comparison_items(term, links),
        },
        {
            "id": "recall-drill",
            "type": "steps",
            "title": "Recall Drill",
            "steps": build_recall_steps(term, taxonomy, links),
        },
        {
            "id": "at-a-glance",
            "type": "table",
            "title": "At a Glance",
            "rows": build_at_a_glance_rows(term, summary, taxonomy, links),
        },
        {
            "id": "visual-summary",
            "type": "callout",
            "title": "Visual Summary",
            "tone": "info",
            "body": family_study_note(term, taxonomy),
        },
        build_curriculum_map_block(launch_contract),
        build_structure_expansion_block(structure_registry),
        build_comparison_block(term, summary, taxonomy, links),
        build_diagram_block(term, taxonomy, links),
        build_faq_block(term, summary, taxonomy, links),
        build_quiz_block(term, taxonomy, links),
    ]

    deep_dive = build_deep_dive_block(term, summary, taxonomy, links)
    if deep_dive and term_content_tier(term, taxonomy, definition, links) == "featured":
        blocks.append(deep_dive)

    if definition:
        blocks.append(
            {
                "id": "source-definition",
                "type": "markdown",
                "title": "Source Definition Snippet",
                "body": definition["body"],
            }
        )

    return [block for block in blocks if block is not None]


def infer_title_aliases(title: str) -> list[str]:
    aliases: list[str] = []
    match = PAREN_ALIAS_RE.match(normalize_label(title))
    if not match:
        return aliases

    base = normalize_label(match.group("base"))
    paren = normalize_label(match.group("paren"))
    if base and base.lower() != title.lower():
        aliases.append(base)
    if paren:
        aliases.append(paren)
    return aliases


def build_term_record(
    term: dict,
    taxonomy_lookup: dict[str, dict],
    definition_lookup: dict[str, dict],
    used_slugs: Counter[str],
    launch_contract: dict,
    structure_registry: dict,
    auto_rules: list[tuple[set[str], str, str, int]] | None = None,
    taxonomy_registry: dict[str, dict] | None = None,
) -> dict:
    slug = resolve_slug(slugify(term["title"]), used_slugs)

    taxonomy_source = "none"
    taxonomy = None

    # Tier 1 — authoritative editorial registry (curated decisions of record).
    if taxonomy_registry:
        registry_candidates = [taxonomy_registry[key] for key in term["titleKeys"] if key in taxonomy_registry]
        if registry_candidates:
            taxonomy = max(
                registry_candidates,
                key=lambda row: (bool(row["category"]), bool(row["subCategory"])),
                default=None,
            )
            if taxonomy and taxonomy["category"]:
                taxonomy_source = "registry"

    # Tier 2 — workbook taxonomy (columns N/O/P of the glossary sheet).
    if not (taxonomy and taxonomy["category"]):
        workbook_candidates = [taxonomy_lookup[key] for key in term["titleKeys"] if key in taxonomy_lookup]
        taxonomy = max(
            workbook_candidates,
            key=lambda row: (bool(row["category"]), bool(row["subCategory"]), -row["rowNumber"]),
            default=None,
        )
        if taxonomy and taxonomy["category"]:
            taxonomy_source = "workbook"

    # Tier 3 — keyword auto-classification (fallback for genuinely new terms).
    was_auto_classified = False
    if not (taxonomy and taxonomy["category"]):
        auto_category, auto_subcategory, classified = auto_classify_term(
            [term["title"], *term.get("variantTitles", [])],
            taxonomy_lookup,
            auto_rules or [],
        )
        if classified:
            taxonomy = {
                "term": term["title"],
                "category": auto_category,
                "subCategory": auto_subcategory,
                "rowNumber": -1,
            }
            was_auto_classified = True
            taxonomy_source = "auto"

    # Tier 4 — study-family taxonomy mapping (metadata-driven fallback).
    # Uses infer_study_family() output to map broad study families
    # (Evaluation, Ethics & Governance, Statistics, Similarity &
    # Deduplication) into concrete taxonomy categories when no other
    # tier has produced a classification. Zero false-positive risk
    # because the family keywords are narrow and domain-specific.
    if not (taxonomy and taxonomy["category"]):
        family = infer_study_family(term, taxonomy)
        if family in STUDY_FAMILY_TAXONOMY_MAP:
            cat, sub = STUDY_FAMILY_TAXONOMY_MAP[family]
            taxonomy = {
                "term": term["title"],
                "category": cat,
                "subCategory": sub,
                "rowNumber": -1,
            }
            taxonomy_source = "study_family"

    definition_candidates = [definition_lookup[key] for key in term["titleKeys"] if key in definition_lookup]
    definition = max(
        definition_candidates,
        key=lambda row: (len(row["body"]), -row["rowNumber"]),
        default=None,
    )
    links = {
        "prerequisites": [],
        "related": [],
        "alternatives": [],
        "next": [],
    }
    summary = build_summary(term, definition, taxonomy, links)
    editorial_tier = term_content_tier(term, taxonomy, definition, links)

    duplicate_rows = term.get("duplicateRows", [])
    source_rows = term.get("inventoryRows", [term["rowNumber"]])

    return {
        "id": slug,
        "slug": slug,
        "title": term["title"],
        "aliases": sorted(
            {
                alias
                for alias in [*term["variantTitles"], *infer_title_aliases(term["title"])]
                if normalize_key(alias) and normalize_key(alias) != term["titleKey"]
            },
            key=str.lower,
        ),
        "summary": summary,
        "taxonomy": {
            "topic": taxonomy["term"] if taxonomy else term["title"],
            "category": taxonomy["category"] if taxonomy else "",
            "subCategory": taxonomy["subCategory"] if taxonomy else "",
            "tags": [
                slugify(value)
                for value in [
                    taxonomy["category"] if taxonomy else "",
                    taxonomy["subCategory"] if taxonomy else "",
                ]
                if value
            ],
        },
        "links": {
            **links,
        },
        "blocks": build_blocks(
            term,
            summary,
            definition,
            taxonomy,
            links,
            launch_contract,
            structure_registry,
            editorial_tier,
        ),
        "metadata": {
            "difficulty": "",
            "maturity": "source-imported",
            "studyFamily": infer_study_family(term, taxonomy),
            "editorialTier": editorial_tier,
            "autoClassified": was_auto_classified,
            "taxonomySource": taxonomy_source,
        },
        "artifact": {
            "shardId": resolve_shard_id(slug),
        },
        "graphTokens": title_tokens_for_graph(term["title"]),
        "source": {
            "glossaryWorkbook": {
                "file": "data_glossary.xlsx",
                "sheetName": "main",
                "inventoryRows": source_rows,
                "inventoryCells": term.get("inventoryCells", []),
                "inventoryColumns": term.get("inventoryColumns", []),
                "sourceCells": term.get("sourceCells", []),
                "taxonomyRow": taxonomy["rowNumber"] if taxonomy else None,
                "definitionRow": definition["rowNumber"] if definition else None,
                "definitionBody": definition["body"] if definition else None,
            }
        },
    }


def build_taxonomy_tree(term_records: list[dict]) -> dict:
    tree: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))

    for term in term_records:
        category = term["taxonomy"]["category"] or "Unclassified"
        sub_category = term["taxonomy"]["subCategory"] or "Unclassified"
        tree[category][sub_category].append({"slug": term["slug"], "title": term["title"]})

    categories = []
    for category_name in sorted(tree):
        subcategories = []
        for subcategory_name in sorted(tree[category_name]):
            members = sorted(tree[category_name][subcategory_name], key=lambda item: item["title"].lower())
            subcategories.append(
                {
                    "name": subcategory_name,
                    "termCount": len(members),
                    "terms": members,
                }
            )
        categories.append({"name": category_name, "subcategories": subcategories})

    return {"categories": categories}


def build_search_index(term_records: list[dict]) -> list[dict]:
    rows = []
    for term in sorted(term_records, key=lambda item: item["title"].lower()):
        rows.append(
            {
                "slug": term["slug"],
                "title": term["title"],
                "aliases": term["aliases"],
                "category": term["taxonomy"]["category"],
                "subCategory": term["taxonomy"]["subCategory"],
                "searchText": " | ".join(
                    part
                    for part in [
                        term["title"],
                        *term["aliases"],
                        term["summary"],
                        term["taxonomy"]["category"],
                        term["taxonomy"]["subCategory"],
                    ]
                    if part
                ),
            }
        )
    return rows


def build_launch_contract(structure_registry: dict) -> dict:
    launch_block_ids = [
        "overview",
        "taxonomy",
        "connections",
        "study-prompts",
        "why-it-matters",
        "comparison-notes",
        "recall-drill",
    ]

    launch_section_map = [
        {
            "section": "Introduction",
            "status": "direct",
            "runtimeBlocks": ["overview", "taxonomy"],
            "note": "The opening definition, context, and classification are always visible.",
        },
        {
            "section": "Prerequisites",
            "status": "direct",
            "runtimeBlocks": ["study-prompts", "connections"],
            "note": "Prerequisite framing is handled by the study sequence and concept graph.",
        },
        {
            "section": "Theoretical Concepts",
            "status": "partial",
            "runtimeBlocks": ["overview", "why-it-matters", "recall-drill"],
            "note": "Core intuition is present, while deeper math-heavy expansion remains editorial backlog.",
        },
        {
            "section": "How It Works",
            "status": "direct",
            "runtimeBlocks": ["study-prompts", "connections"],
            "note": "The step sequence and relation map carry the process explanation.",
        },
        {
            "section": "Variants or Extensions",
            "status": "direct",
            "runtimeBlocks": ["connections", "comparison-notes"],
            "note": "Alternatives and neighboring labels are surfaced through graph and comparison blocks.",
        },
        {
            "section": "Applications",
            "status": "direct",
            "runtimeBlocks": ["overview", "why-it-matters"],
            "note": "The summary and editorial note show relevance before deeper examples are added.",
        },
        {
            "section": "Implementation",
            "status": "direct",
            "runtimeBlocks": ["study-prompts", "recall-drill"],
            "note": "Implementation thinking is supported by the guided sequence and recall practice.",
        },
        {
            "section": "Evaluation and Metrics",
            "status": "direct",
            "runtimeBlocks": ["comparison-notes", "recall-drill"],
            "note": "Tradeoffs, comparison, and recall drills anchor evaluation thinking.",
        },
        {
            "section": "Advantages and Disadvantages",
            "status": "direct",
            "runtimeBlocks": ["comparison-notes", "why-it-matters"],
            "note": "The current blocks already surface tradeoffs and framing.",
        },
        {
            "section": "Ethics and Responsible AI",
            "status": "direct",
            "runtimeBlocks": ["why-it-matters", "comparison-notes"],
            "note": "Responsible-AI framing is visible in the editorial layer and comparison notes.",
        },
        {
            "section": "Related Concepts",
            "status": "direct",
            "runtimeBlocks": ["connections"],
            "note": "The concept graph is already a first-class runtime block.",
        },
    ]

    return {
        "sheetName": structure_registry["sheetName"],
        "launchSectionCount": len(launch_section_map),
        "launchBlockIds": launch_block_ids,
        "launchSections": launch_section_map,
        "policy": "Launch runtime stays compact; structure workbook sections beyond this map remain editorial-expansion or backlog unless promoted later.",
        "sourceStructureCounts": {
            "launch-runtime": structure_registry["layerCounts"]["launch-runtime"],
            "editorial-expansion": structure_registry["layerCounts"]["editorial-expansion"],
            "backlog": structure_registry["layerCounts"]["backlog"],
        },
    }


def build_catalog_index(term_records: list[dict]) -> list[dict]:
    return [
        {
            "id": term["id"],
            "slug": term["slug"],
            "title": term["title"],
            "aliases": term["aliases"],
            "summary": term["summary"],
            "taxonomy": term["taxonomy"],
            "links": term["links"],
            "metadata": term["metadata"],
            "artifact": term["artifact"],
        }
        for term in sorted(term_records, key=lambda item: item["title"].lower())
    ]


def load_path_sequences(sequences_path: Path | None) -> dict:
    """Load editorial path sequences from editorial/path-sequences.json.

    Returns a dict mapping path_slug -> {steps, category, subCategory}.
    Each step has: {slug, title, stage, why}.
    Returns {} when no path is given or the file is absent.
    """
    if sequences_path is None:
        return {}
    path = Path(sequences_path)
    if not path.is_file():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(payload, dict):
            return payload
    except (json.JSONDecodeError, OSError):
        pass
    return {}


def classify_path_stage(index: int, total: int) -> str:
    if index == 0:
        return "start"
    if total <= 4:
        return "build" if index < total - 1 else "advance"
    if index < 2:
        return "start"
    if index < min(5, total - 1):
        return "build"
    return "advance"


def term_path_score(term: dict) -> tuple[int, int, int, int, int]:
    link_count = sum(len(term["links"][key]) for key in ["prerequisites", "related", "alternatives", "next"])
    has_definition = 1 if term["source"]["glossaryWorkbook"]["definitionRow"] else 0
    has_aliases = len(term["aliases"])
    taxonomy_depth = int(bool(term["taxonomy"]["category"])) + int(bool(term["taxonomy"]["subCategory"]))
    return (
        has_definition,
        link_count,
        taxonomy_depth,
        has_aliases,
        -len(term["title"]),
    )


def build_learning_paths(term_records: list[dict], path_sequences: dict | None = None) -> tuple[list[dict], dict[str, dict]]:
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for term in term_records:
        category = term["taxonomy"]["category"]
        subcategory = term["taxonomy"]["subCategory"]
        if not category or not subcategory:
            continue
        grouped[(category, subcategory)].append(term)

    path_summaries: list[dict] = []
    path_details: dict[str, dict] = {}
    sequences = path_sequences if path_sequences else {}

    for (category, subcategory), terms in sorted(grouped.items()):
        if len(terms) < 3:
            continue

        path_slug = slugify(f"{category} {subcategory}")

        # Check for a manually sequenced path override
        manual = sequences.get(path_slug)
        if manual and manual.get("steps"):
            term_by_slug: dict[str, dict] = {t["slug"]: t for t in terms}
            manual_steps = manual["steps"]
            steps = []
            for step_def in manual_steps:
                slug = step_def["slug"]
                term = term_by_slug.get(slug)
                if not term:
                    continue
                steps.append(
                    {
                        "slug": term["slug"],
                        "title": term["title"],
                        "summary": term["summary"],
                        "stage": step_def.get("stage", "build"),
                        "whyIncluded": step_def.get("why", ""),
                    }
                )

            if len(steps) >= 2:
                description = (
                    f"Build fluency in {subcategory} inside {category} by moving from foundational terms "
                    "into adjacent methods, variants, and extensions."
                )
                summary = {
                    "slug": path_slug,
                    "title": f"{subcategory} Path",
                    "description": description,
                    "category": category,
                    "subCategory": subcategory,
                    "termCount": len(terms),
                    "featuredTermSlugs": [step["slug"] for step in steps[:4]],
                }
                detail = {
                    **summary,
                    "featuredTermTitles": [step["title"] for step in steps[:4]],
                    "steps": steps,
                }
                path_summaries.append(summary)
                path_details[path_slug] = detail
                continue

        # Auto-generate path (fallback when no manual sequence exists)
        ordered = sorted(
            terms,
            key=lambda term: (
                *term_path_score(term),
                term["title"].lower(),
            ),
            reverse=True,
        )
        selected = ordered[:8]
        steps = []
        for index, term in enumerate(selected):
            stage = classify_path_stage(index, len(selected))
            if stage == "start":
                why = "Start here to anchor the vocabulary and core intuition for this area."
            elif stage == "build":
                why = "Use this after the foundations to connect patterns, variants, and system behavior."
            else:
                why = "Treat this as an advanced extension once the earlier terms feel natural."

            steps.append(
                {
                    "slug": term["slug"],
                    "title": term["title"],
                    "summary": term["summary"],
                    "stage": stage,
                    "whyIncluded": why,
                }
            )

        description = (
            f"Build fluency in {subcategory} inside {category} by moving from foundational terms "
            "into adjacent methods, variants, and extensions."
        )
        summary = {
            "slug": path_slug,
            "title": f"{subcategory} Path",
            "description": description,
            "category": category,
            "subCategory": subcategory,
            "termCount": len(terms),
            "featuredTermSlugs": [step["slug"] for step in steps[:4]],
        }
        detail = {
            **summary,
            "featuredTermTitles": [step["title"] for step in steps[:4]],
            "steps": steps,
        }
        path_summaries.append(summary)
        path_details[path_slug] = detail

    path_summaries.sort(key=lambda item: (-item["termCount"], item["category"].lower(), item["subCategory"].lower()))
    return path_summaries, path_details


def build_term_shards(term_records: list[dict]) -> dict[str, list[dict]]:
    shards: dict[str, list[dict]] = defaultdict(list)
    for term in term_records:
        shards[term["artifact"]["shardId"]].append(term)

    return {
        shard_id: sorted(rows, key=lambda item: item["title"].lower())
        for shard_id, rows in sorted(shards.items())
    }


def dedupe_titles(values: list[str], limit: int) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for value in values:
        key = normalize_key(value)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(value)
        if len(deduped) >= limit:
            break
    return deduped


def rank_peer(term: dict, peer: dict) -> tuple[int, int, int, int]:
    current_tokens = set(term["graphTokens"])
    peer_tokens = set(peer["graphTokens"])
    shared = len(current_tokens & peer_tokens)
    category_bonus = 1 if peer["taxonomy"]["category"] else 0
    subset_bonus = 1 if peer_tokens and peer_tokens.issubset(current_tokens) else 0
    return (
        shared,
        subset_bonus,
        category_bonus,
        -abs(len(set(peer["graphTokens"])) - len(set(term["graphTokens"]))),
    )


def lexical_peers(term: dict, token_index: dict[str, list[dict]]) -> list[dict]:
    rows: dict[str, dict] = {}
    for token in term["graphTokens"]:
        for peer in token_index.get(token, []):
            if peer["slug"] == term["slug"]:
                continue
            rows[peer["slug"]] = peer
    return sorted(rows.values(), key=lambda peer: (*rank_peer(term, peer), peer["title"].lower()), reverse=True)


def choose_related_titles(term: dict, peer_groups: list[list[dict]], token_index: dict[str, list[dict]], limit: int = 4) -> list[str]:
    titles: list[str] = []

    for peers in peer_groups:
        ranked = sorted(
            [peer for peer in peers if peer["slug"] != term["slug"]],
            key=lambda peer: (*rank_peer(term, peer), peer["title"].lower()),
            reverse=True,
        )
        titles.extend(peer["title"] for peer in ranked)
        titles = dedupe_titles(titles, limit)
        if len(titles) >= limit:
            return titles

    lexical_ranked = lexical_peers(term, token_index)
    titles.extend(peer["title"] for peer in lexical_ranked)
    return dedupe_titles(titles, limit)


def choose_prerequisite_titles(term: dict, peer_groups: list[list[dict]], token_index: dict[str, list[dict]], limit: int = 3) -> list[str]:
    current_tokens = set(term["graphTokens"])
    titles: list[str] = []

    def subset_titles(peers: list[dict]) -> list[str]:
        candidates = []
        for peer in peers:
            if peer["slug"] == term["slug"]:
                continue
            peer_tokens = set(peer["graphTokens"])
            if not peer_tokens or len(peer_tokens) >= len(current_tokens):
                continue
            if peer_tokens.issubset(current_tokens):
                candidates.append(peer)
        candidates.sort(key=lambda peer: (-len(set(peer["graphTokens"])), len(peer["title"]), peer["title"].lower()))
        return [peer["title"] for peer in candidates]

    for peers in peer_groups:
        titles.extend(subset_titles(peers))
        titles = dedupe_titles(titles, limit)
        if len(titles) >= limit:
            return titles

    titles.extend(subset_titles(lexical_peers(term, token_index)))
    return dedupe_titles(titles, limit)


def choose_next_titles(term: dict, peer_groups: list[list[dict]], token_index: dict[str, list[dict]], limit: int = 3) -> list[str]:
    current_tokens = set(term["graphTokens"])
    titles: list[str] = []

    def superset_titles(peers: list[dict]) -> list[str]:
        candidates = []
        for peer in peers:
            if peer["slug"] == term["slug"]:
                continue
            peer_tokens = set(peer["graphTokens"])
            if len(peer_tokens) <= len(current_tokens):
                continue
            if current_tokens and current_tokens.issubset(peer_tokens):
                candidates.append(peer)
        candidates.sort(key=lambda peer: (len(set(peer["graphTokens"])), len(peer["title"]), peer["title"].lower()))
        return [peer["title"] for peer in candidates]

    for peers in peer_groups:
        titles.extend(superset_titles(peers))
        titles = dedupe_titles(titles, limit)
        if len(titles) >= limit:
            return titles

    titles.extend(superset_titles(lexical_peers(term, token_index)))
    return dedupe_titles(titles, limit)


def choose_alternative_titles(term: dict, semantic_groups: dict[str, list[dict]], limit: int = 4) -> list[str]:
    alternatives = [
        peer["title"]
        for peer in semantic_groups.get(semantic_key(term["title"]), [])
        if peer["slug"] != term["slug"]
    ]
    return alternatives[:limit]


def enrich_term_links(term_records: list[dict]) -> None:
    by_subcategory: dict[tuple[str, str], list[dict]] = defaultdict(list)
    by_category: dict[str, list[dict]] = defaultdict(list)
    semantic_groups: dict[str, list[dict]] = defaultdict(list)
    token_index: dict[str, list[dict]] = defaultdict(list)

    for term in term_records:
        category = normalize_key(term["taxonomy"]["category"])
        subcategory = normalize_key(term["taxonomy"]["subCategory"])
        if category and subcategory:
            by_subcategory[(category, subcategory)].append(term)
        if category:
            by_category[category].append(term)
        semantic_groups[semantic_key(term["title"])].append(term)
        for token in term["graphTokens"]:
            token_index[token].append(term)

    for term in term_records:
        category = normalize_key(term["taxonomy"]["category"])
        subcategory = normalize_key(term["taxonomy"]["subCategory"])
        peer_groups = []
        if category and subcategory:
            peer_groups.append(by_subcategory.get((category, subcategory), []))
        if category:
            peer_groups.append(by_category.get(category, []))

        term["links"] = {
            "prerequisites": choose_prerequisite_titles(term, peer_groups, token_index),
            "related": choose_related_titles(term, peer_groups, token_index),
            "alternatives": choose_alternative_titles(term, semantic_groups),
            "next": choose_next_titles(term, peer_groups, token_index),
        }


def finalize_term_content(term_records: list[dict], launch_contract: dict, structure_registry: dict) -> None:
    for term in term_records:
        glossary_source = term.get("source", {}).get("glossaryWorkbook", {})
        definition_body = glossary_source.get("definitionBody")
        definition = {"body": definition_body} if definition_body else None
        term["metadata"]["editorialTier"] = term_content_tier(term, term.get("taxonomy"), definition, term.get("links", {}))
        term["summary"] = build_summary(term, definition, term["taxonomy"], term["links"])
        term["blocks"] = build_blocks(
            term,
            term["summary"],
            definition,
            term["taxonomy"],
            term["links"],
            launch_contract,
            structure_registry,
            term["metadata"]["editorialTier"],
        )


def strip_internal_fields(term_records: list[dict]) -> None:
    for term in term_records:
        term.pop("graphTokens", None)


def build_shard_manifest(term_shards: dict[str, list[dict]]) -> dict:
    return {
        "shardCount": len(term_shards),
        "totalTerms": sum(len(rows) for rows in term_shards.values()),
        "shards": [
            {
                "id": shard_id,
                "termCount": len(rows),
                "firstSlug": rows[0]["slug"] if rows else None,
                "lastSlug": rows[-1]["slug"] if rows else None,
            }
            for shard_id, rows in term_shards.items()
        ],
    }


def build_content_audit(
    term_records: list[dict],
    source_inventory: list[dict],
    canonicalization_groups: list[dict],
) -> dict:
    source_title_keys = {term["titleKey"] for term in source_inventory}
    required_block_totals = {block_id: 0 for block_id in EXPECTED_BLOCK_SEQUENCE}
    issue_counts: Counter[str] = Counter()
    issue_samples: dict[str, list[dict]] = defaultdict(list)
    coverage = Counter()
    content_tier_counts = Counter()
    high_severity_count = 0

    def add_issue(issue_type: str, severity: str, term: dict, detail: str) -> None:
        nonlocal high_severity_count
        issue_counts[issue_type] += 1
        if severity == "high":
            high_severity_count += 1
        samples = issue_samples[issue_type]
        if len(samples) < 10:
            samples.append(
                {
                    "slug": term["slug"],
                    "title": term["title"],
                    "severity": severity,
                    "detail": detail,
                }
            )

    for term in term_records:
        title_key = normalize_key(term["title"])
        blocks = term.get("blocks", [])
        block_ids = [block.get("id") for block in blocks]
        source = term.get("source", {}).get("glossaryWorkbook", {})
        aliases = term.get("aliases", [])
        summary = term.get("summary", "")

        coverage["terms"] += 1
        if term.get("taxonomy", {}).get("category"):
            coverage["taxonomy"] += 1
            taxonomy_source = term.get("metadata", {}).get("taxonomySource", "none")
            if taxonomy_source == "registry":
                coverage["registryClassifiedTerms"] += 1
            elif taxonomy_source == "workbook":
                coverage["workbookClassifiedTerms"] += 1
            elif taxonomy_source == "auto":
                coverage["autoClassifiedTerms"] += 1
        if not term.get("taxonomy", {}).get("category"):
            coverage["unclassifiedTerms"] += 1
        if source.get("definitionRow"):
            coverage["definition"] += 1
        if term.get("metadata", {}).get("studyFamily"):
            coverage["studyFamily"] += 1
        if term.get("metadata", {}).get("editorialTier"):
            content_tier_counts[str(term["metadata"]["editorialTier"])] += 1
        if block_ids:
            coverage["blocks"] += 1
        if source.get("definitionRow") and any(block.get("id") == "source-definition" for block in blocks):
            coverage["sourceDefinitionBlock"] += 1

        if title_key not in source_title_keys:
            add_issue("title_missing_from_source_inventory", "high", term, "Published title is not backed by a source inventory row.")

        if not source.get("inventoryRows"):
            add_issue("missing_inventory_rows", "high", term, "Term has no recorded source inventory rows.")

        if block_ids[: len(EXPECTED_BLOCK_SEQUENCE)] != list(EXPECTED_BLOCK_SEQUENCE):
            add_issue(
                "required_block_order_mismatch",
                "high",
                term,
                f"Expected the first four blocks to be {list(EXPECTED_BLOCK_SEQUENCE)}.",
            )
        for block_id in EXPECTED_BLOCK_SEQUENCE:
            if block_id not in block_ids:
                add_issue("required_block_missing", "high", term, f"Missing required block '{block_id}'.")
            else:
                required_block_totals[block_id] += 1

        has_definition_block = any(block.get("id") == "source-definition" for block in blocks)
        has_source_definition = bool(source.get("definitionRow"))
        if has_definition_block != has_source_definition:
            add_issue(
                "source_definition_block_mismatch",
                "high",
                term,
                "Source-definition block parity does not match source definition coverage.",
            )
        elif has_definition_block:
            source_definition_body = normalize_label(source.get("definitionBody"))
            block_definition = normalize_label(next(block.get("body") for block in blocks if block.get("id") == "source-definition"))
            if source_definition_body != block_definition:
                add_issue(
                    "source_definition_body_mismatch",
                    "high",
                    term,
                    "Source-definition block body differs from the workbook definition body.",
                )

        if not summary:
            add_issue("summary_missing", "high", term, "Summary is missing entirely.")
        elif len(summary) < 60:
            coverage["conciseSummary"] += 1

        if any(
            marker in summary.lower()
            for marker in (
                "included in the canonical glossary",
                "use the linked concepts, study prompts, and your own notes",
            )
        ):
            add_issue(
                "fallback_summary",
                "low",
                term,
                "Summary used the sparse-source fallback copy instead of a taxonomy or definition-backed sentence.",
            )

        if normalize_key(term["title"]) in {"term", "terms", "topic"}:
            add_issue("generic_title", "high", term, "Title is too generic to be a trustworthy learning entry.")

        if not re.search(r"[a-zA-Z0-9]", term["title"]):
            add_issue("non_alphanumeric_title", "medium", term, "Title does not contain a standard alphanumeric token.")

        if len(aliases) != len({normalize_key(alias) for alias in aliases}):
            add_issue("duplicate_aliases", "medium", term, "Aliases contain normalized duplicates.")

        if term.get("links", {}).get("prerequisites") and any(
            normalize_key(link) == title_key for link in term["links"]["prerequisites"]
        ):
            add_issue("self_referential_prerequisite", "high", term, "Term lists itself as a prerequisite.")
        if term.get("links", {}).get("related") and any(normalize_key(link) == title_key for link in term["links"]["related"]):
            add_issue("self_referential_related", "high", term, "Term lists itself as a related concept.")
        if term.get("links", {}).get("alternatives") and any(
            normalize_key(link) == title_key for link in term["links"]["alternatives"]
        ):
            add_issue("self_referential_alternative", "high", term, "Term lists itself as an alternative.")
        if term.get("links", {}).get("next") and any(normalize_key(link) == title_key for link in term["links"]["next"]):
            add_issue("self_referential_next", "high", term, "Term lists itself as a next step.")

    return {
        "auditedAt": datetime.now(timezone.utc).isoformat(),
        "status": "pass" if high_severity_count == 0 else "fail",
        "termCount": len(term_records),
        "coverage": {
            "taxonomyTerms": coverage["taxonomy"],
            "taxonomyCoverageRatio": round(coverage["taxonomy"] / len(term_records), 6) if term_records else 0,
            "registryClassifiedTerms": coverage.get("registryClassifiedTerms", 0),
            "workbookClassifiedTerms": coverage.get("workbookClassifiedTerms", 0),
            "autoClassifiedTerms": coverage.get("autoClassifiedTerms", 0),
            "unclassifiedTerms": coverage.get("unclassifiedTerms", 0),
            "definitionTerms": coverage["definition"],
            "definitionCoverageRatio": round(coverage["definition"] / len(term_records), 6) if term_records else 0,
            "studyFamilyTerms": coverage["studyFamily"],
            "studyFamilyCoverageRatio": round(coverage["studyFamily"] / len(term_records), 6) if term_records else 0,
            "sourceDefinitionBlocks": coverage["sourceDefinitionBlock"],
            "blockCoverage": {block_id: required_block_totals[block_id] for block_id in EXPECTED_BLOCK_SEQUENCE},
        },
        "contentTierCounts": dict(sorted(content_tier_counts.items())),
        "qualityChecks": {
            "requiredBlockSequence": list(EXPECTED_BLOCK_SEQUENCE),
            "sourceInventoryBackedTitles": high_severity_count == 0,
            "highSeverityIssueCount": high_severity_count,
            "issueCount": sum(issue_counts.values()),
        },
        "issueCounts": dict(sorted(issue_counts.items())),
        "issueSamples": {key: value for key, value in sorted(issue_samples.items())},
        "canonicalization": {
            "groups": len(canonicalization_groups),
            "rowsMerged": sum(group["mergedRowCount"] - 1 for group in canonicalization_groups),
        },
    }


def build_published_manifest(
    source_inventory: list[dict],
    term_records: list[dict],
    path_summaries: list[dict],
    term_shards: dict[str, list[dict]],
    report_payload: dict,
    content_audit: dict,
    structure_registry: dict,
    launch_contract: dict,
    content_depth: dict,
    source_hashes: dict[str, str],
    artifact_hash_report_hash: str,
) -> dict:
    return {
        "schemaVersion": PUBLISHED_SCHEMA_VERSION,
        "contentVersion": CONTENT_VERSION,
        "builderVersion": BUILDER_VERSION,
        "publishedAt": content_audit["auditedAt"],
        "status": content_audit["status"],
        "sourceInventoryTermCount": len(source_inventory),
        "termCount": len(term_records),
        "pathCount": len(path_summaries),
        "shardCount": len(term_shards),
        "coverage": content_audit["coverage"],
        "qualityChecks": content_audit["qualityChecks"],
        "canonicalization": content_audit["canonicalization"],
        "contentTierCounts": content_audit["contentTierCounts"],
        "structureLayerCounts": structure_registry["layerCounts"],
        "editorialFieldCount": report_payload["editorialFieldCount"],
        "structureSectionCount": report_payload["structureSectionCount"],
        "launchStructureSections": report_payload["launchStructureSections"],
        "editorialStructureSections": report_payload["editorialStructureSections"],
        "launchSectionCount": launch_contract["launchSectionCount"],
        "launchBlockIds": launch_contract["launchBlockIds"],
        "launchSections": launch_contract["launchSections"],
        "contentDepth": content_depth,
        "sourceHashes": source_hashes,
        "artifactHashReport": "reports/artifact-hashes.json",
        "artifactHashReportHash": artifact_hash_report_hash,
        "artifacts": {
            "termsIndex": "terms/index.json",
            "termsManifest": "terms/manifest.json",
            "pathsIndex": "paths/index.json",
            "taxonomyTree": "taxonomy/category-tree.json",
            "searchIndex": "search/search-index.json",
            "importReport": "reports/import-report.json",
            "duplicateGroups": "reports/duplicate-groups.json",
            "canonicalizationGroups": "reports/canonicalization-groups.json",
            "contentAudit": "reports/content-audit.json",
            "artifactHashes": "reports/artifact-hashes.json",
            "structureRegistry": "editorial/structure-registry.json",
            "launchContract": "editorial/launch-contract.json",
        },
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--glossary-workbook", required=True, type=Path)
    parser.add_argument("--structure-workbook", required=True, type=Path)
    parser.add_argument("--out-dir", required=True, type=Path)
    parser.add_argument(
        "--taxonomy-registry",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "data" / "taxonomy-registry.json",
        help="Authoritative editorial taxonomy registry (default: data/taxonomy-registry.json). "
             "Generate with tools/extract_taxonomy_registry.py. Pass a non-existent path to disable.",
    )
    parser.add_argument(
        "--path-sequences",
        type=Path,
        default=None,
        help="Optional editorial path sequence JSON. Defaults to the existing output directory editorial/path-sequences.json when present.",
    )
    parser.add_argument(
        "--coverage-threshold",
        type=float,
        default=TAXONOMY_COVERAGE_MINIMUM,
        help="Override the taxonomy coverage minimum (default: %(default)s). "
             "Use 0 to disable the gate entirely for dev builds.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    final_out_dir = args.out_dir
    staging_root = final_out_dir.parent
    staging_root.mkdir(parents=True, exist_ok=True)
    args.out_dir = Path(tempfile.mkdtemp(prefix=f".{final_out_dir.name}.staging.", dir=staging_root))

    source_inventory = extract_term_inventory(args.glossary_workbook)
    inventory, canonicalization_groups = merge_semantic_variants(source_inventory)
    taxonomy_lookup = extract_taxonomy_lookup(args.glossary_workbook)
    definition_lookup = extract_definition_lookup(args.glossary_workbook)
    structure_registry = extract_structure_registry(args.structure_workbook)
    launch_contract = build_launch_contract(structure_registry)
    taxonomy_registry = load_taxonomy_registry(args.taxonomy_registry)
    source_path_sequences = Path("content/source/path-sequences.json")
    path_sequences_path = args.path_sequences or (source_path_sequences if source_path_sequences.exists() else final_out_dir / "editorial" / "path-sequences.json")
    path_sequences = load_path_sequences(path_sequences_path)

    used_slugs: Counter[str] = Counter()
    term_records = [
        build_term_record(
            term, taxonomy_lookup, definition_lookup, used_slugs, launch_contract, structure_registry,
            auto_rules=AUTO_CLASSIFICATION_RULES,
            taxonomy_registry=taxonomy_registry,
        )
        for term in inventory
    ]
    enrich_term_links(term_records)
    finalize_term_content(term_records, launch_contract, structure_registry)
    strip_internal_fields(term_records)
    term_shards = build_term_shards(term_records)
    path_summaries, path_details = build_learning_paths(term_records, path_sequences)
    content_audit = build_content_audit(term_records, source_inventory, canonicalization_groups)
    block_counts = Counter()
    for term in term_records:
        for block in term.get("blocks", []):
            block_id = block.get("id")
            if block_id:
                block_counts[block_id] += 1
    content_depth = {
        "blockCounts": dict(sorted(block_counts.items())),
        "featuredTermCount": sum(1 for term in term_records if term["metadata"]["editorialTier"] == "featured"),
        "sourceDefinitionCount": sum(1 for term in term_records if term["source"]["glossaryWorkbook"]["definitionRow"]),
    }

    for shard_id, rows in term_shards.items():
        write_json(
            args.out_dir / "terms" / "shards" / f"{shard_id}.json",
            {
                "id": shard_id,
                "termCount": len(rows),
                "terms": rows,
            },
        )

    taxonomy_matches = sum(1 for term in term_records if term["taxonomy"]["category"])
    registry_classified = sum(1 for term in term_records if term["metadata"].get("taxonomySource") == "registry")
    workbook_classified = sum(1 for term in term_records if term["metadata"].get("taxonomySource") == "workbook")
    auto_classified = sum(1 for term in term_records if term["metadata"].get("taxonomySource") == "auto")
    study_family_classified = sum(1 for term in term_records if term["metadata"].get("taxonomySource") == "study_family")
    unclassified = sum(1 for term in term_records if not term["taxonomy"]["category"])
    definition_matches = sum(1 for term in term_records if term["source"]["glossaryWorkbook"]["definitionRow"])
    duplicate_groups = [
        {
            "title": term["title"],
            "slug": term["slug"],
            "inventoryRows": term["source"]["glossaryWorkbook"]["inventoryRows"],
            "duplicateRowCount": max(0, len(term["source"]["glossaryWorkbook"]["inventoryRows"]) - 1),
        }
        for term in term_records
        if len(term["source"]["glossaryWorkbook"]["inventoryRows"]) > 1
    ]

    report_payload = {
        "sourceInventoryTermCount": len(source_inventory),
        "termCount": len(term_records),
        "taxonomyMatches": taxonomy_matches,
        "taxonomyCoverageRatio": round(taxonomy_matches / len(term_records), 6) if term_records else 0,
        "registryClassifiedTerms": registry_classified,
        "workbookClassifiedTerms": workbook_classified,
        "autoClassifiedTerms": auto_classified,
        "studyFamilyClassifiedTerms": study_family_classified,
        "unclassifiedTerms": unclassified,
        "definitionMatches": definition_matches,
        "definitionCoverageRatio": round(definition_matches / len(term_records), 6) if term_records else 0,
        "duplicateGroups": len(duplicate_groups),
        "slugCollisionCount": sum(1 for count in used_slugs.values() if count > 1),
        "canonicalizationGroups": len(canonicalization_groups),
        "canonicalizationRowsMerged": sum(group["mergedRowCount"] - 1 for group in canonicalization_groups),
        "shardCount": len(term_shards),
        "largestShardTermCount": max((len(rows) for rows in term_shards.values()), default=0),
        "pathCount": len(path_summaries),
        "editorialFieldCount": structure_registry["fieldCount"],
        "structureLayerCounts": structure_registry["layerCounts"],
        "structureSectionCount": len(structure_registry["sectionGroups"]),
        "launchStructureSections": len(structure_registry["launchSections"]),
        "editorialStructureSections": len(structure_registry["editorialSections"]),
    }

    write_json(args.out_dir / "terms" / "index.json", build_catalog_index(term_records))
    write_json(args.out_dir / "terms" / "manifest.json", build_shard_manifest(term_shards))
    # Per-term detail files: the frontend lazy-loads each term by slug
    # (src/content/CatalogContext.tsx). These are the same finalized term
    # records the shards carry, projected one-file-per-term for direct fetch.
    for term in term_records:
        write_json(args.out_dir / "terms" / "by-slug" / f"{term['slug']}.json", term)
    write_json(args.out_dir / "paths" / "index.json", path_summaries)
    for path_slug, detail in path_details.items():
        write_json(args.out_dir / "paths" / "by-slug" / f"{path_slug}.json", detail)
    write_json(args.out_dir / "taxonomy" / "category-tree.json", build_taxonomy_tree(term_records))
    write_json(args.out_dir / "search" / "search-index.json", build_search_index(term_records))
    write_json(args.out_dir / "reports" / "import-report.json", report_payload)
    write_json(args.out_dir / "reports" / "duplicate-groups.json", duplicate_groups)
    write_json(args.out_dir / "reports" / "canonicalization-groups.json", canonicalization_groups)
    write_json(args.out_dir / "reports" / "content-audit.json", content_audit)
    write_json(args.out_dir / "editorial" / "structure-registry.json", structure_registry)
    write_json(args.out_dir / "editorial" / "launch-contract.json", launch_contract)
    if path_sequences:
        if path_sequences_path and path_sequences_path.exists():
            target_path_sequences = args.out_dir / "editorial" / "path-sequences.json"
            target_path_sequences.parent.mkdir(parents=True, exist_ok=True)
            shutil.copyfile(path_sequences_path, target_path_sequences)
        else:
            write_json(args.out_dir / "editorial" / "path-sequences.json", path_sequences)
    source_hashes = {
        "glossaryWorkbook": file_sha256(args.glossary_workbook),
        "structureWorkbook": file_sha256(args.structure_workbook),
    }
    if args.taxonomy_registry and args.taxonomy_registry.exists():
        source_hashes["taxonomyRegistry"] = file_sha256(args.taxonomy_registry)
    emerging_terms_path = Path("content/source/emerging-terms-2024-2026.json")
    if emerging_terms_path.exists():
        source_hashes["emergingTerms"] = file_sha256(emerging_terms_path)
    if path_sequences_path and path_sequences_path.exists():
        source_hashes["pathSequences"] = file_sha256(path_sequences_path)
    artifact_hashes = collect_artifact_hashes(args.out_dir)
    write_json(args.out_dir / "reports" / "artifact-hashes.json", artifact_hashes)
    artifact_hash_report_hash = file_sha256(args.out_dir / "reports" / "artifact-hashes.json")
    write_json(
        args.out_dir / "manifest.json",
        build_published_manifest(
            source_inventory,
            term_records,
            path_summaries,
            term_shards,
            report_payload,
            content_audit,
            structure_registry,
            launch_contract,
            content_depth,
            source_hashes,
            artifact_hash_report_hash,
        ),
    )
    coverage_ratio = content_audit["coverage"]["taxonomyCoverageRatio"]
    if args.coverage_threshold > 0 and coverage_ratio < args.coverage_threshold:
        print(
            f"WARNING: Taxonomy coverage ratio {coverage_ratio:.4f} "
            f"is below minimum threshold {args.coverage_threshold:.2f}. "
            "Add more rules to AUTO_CLASSIFICATION_RULES or extend the workbook taxonomy."
        )

    if content_audit["qualityChecks"]["highSeverityIssueCount"] > 0:
        raise RuntimeError("Content audit failed with high-severity issues; inspect reports/content-audit.json.")

    replace_directory_after_success(args.out_dir, final_out_dir)
    print(f"Published {len(term_records)} terms into {final_out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
